import pandas as pd
import openpyxl
import sys

# Set stdout to utf-8 just in case
sys.stdout.reconfigure(encoding='utf-8')

file_path = r"C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\legacy_excel\REGISTRO HORAS  (1).xlsm"
output_file = r"C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\excel_analysis_result.txt"

try:
    with open(output_file, 'w', encoding='utf-8') as f:
        # Load workbook to read formulas
        wb = openpyxl.load_workbook(file_path, data_only=False)
        
        f.write("--- SHEETS ---\n")
        f.write(str(wb.sheetnames) + "\n")
        
        if "HOJA_BASE" in wb.sheetnames:
            ws = wb["HOJA_BASE"]
            f.write("\n--- HOJA_BASE (BE2:BH20) ---\n")
            
            # Headers for BE1:BH1
            headers = []
            for col in range(57, 61): # BE=57
                cell = ws.cell(row=1, column=col)
                headers.append(f"{cell.coordinate}: {cell.value}")
            f.write("Headers: " + str(headers) + "\n")

            for row in range(2, 21):
                row_data = []
                for col in range(57, 61): 
                    cell = ws.cell(row=row, column=col)
                    val = cell.value
                    if val is None: val = ""
                    row_data.append(f"{cell.coordinate}={val}")
                f.write(f"Row {row}: {', '.join(row_data)}\n")
                
            f.write("\n--- HOJA_BASE (First 20 Columns) ---\n")
            # Check first few columns to see structure
            for row in range(1, 5):
                r = [f"{c.coordinate}={c.value}" for c in ws[row][:20]] 
                f.write(str(r) + "\n")
                
        if "FESTIVOS" in wb.sheetnames:
            f.write("\n--- FESTIVOS ---\n")
            ws_fest = wb["FESTIVOS"]
            for row in ws_fest.iter_rows(min_row=1, max_row=10, values_only=True):
                f.write(str(row) + "\n")

    print("Done writing to file.")
except Exception as e:
    print(f"Error: {e}")
