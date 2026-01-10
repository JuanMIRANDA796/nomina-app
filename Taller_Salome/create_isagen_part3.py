import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("="*80)
print("AÑADIENDO INDICADORES FINANCIEROS Y FLUJOS DE CAJA")
print("="*80)

# Load existing workbook
wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = load_workbook(wb_file)

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
subheader_fill = PatternFill(start_color="7BA0C9", end_color="7BA0C9", fill_type="solid")
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
indicator_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
cashflow_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
bold_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

# ==================== SHEET 7: INDICADORES FINANCIEROS ====================
print("\n📊 Creando hoja 7: Indicadores Financieros...")
ws_ind = wb.create_sheet("7. Indicadores Financieros")
ws_ind.sheet_view.showGridLines = False

# Title
ws_ind['A1'] = 'ISAGEN S.A. E.S.P.'
ws_ind['A1'].font = Font(name='Arial', size=14, bold=True)
ws_ind['A2'] = 'INDICADORES FINANCIEROS CLAVE'
ws_ind['A2'].font = Font(name='Arial', size=12, bold=True)
ws_ind['A3'] = 'Valores en millones de pesos colombianos (COP)'
ws_ind['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_ind.cell(row, 1, 'INDICADOR')
ws_ind.cell(row, 2, '2020')
ws_ind.cell(row, 3, '2021')
ws_ind.cell(row, 4, '2022')
ws_ind.cell(row, 5, '2023')
ws_ind.cell(row, 6, '2024')
ws_ind.cell(row, 7, '2025 (Sep)')
format_header(ws_ind, row, 7)

# Mapping of rows in balance sheet (approximate based on common structure)
# We'll reference specific accounts
current_row = 6

# Section: EBITDA (already calculated)
ws_ind.cell(current_row, 1, 'EBITDA')
ws_ind.cell(current_row, 1).font = bold_font
ws_ind.cell(current_row, 1).fill = total_fill
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # EBITDA is in row 19 of Estado de Resultados
    formula = f"='2. Estado de Resultados'!{col_letter}19"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '#,##0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
    ws_ind.cell(current_row, col_idx).fill = formula_fill
current_row += 2

# Section: Capital de Trabajo
ws_ind.cell(current_row, 1, 'CAPITAL DE TRABAJO NETO OPERATIVO (KTNO)')
ws_ind.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_ind.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_ind.cell(current_row, col).fill = subheader_fill
current_row += 1

# Activos Corrientes Operativos
ws_ind.cell(current_row, 1, 'Activos Corrientes Operativos')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # Total Activos Corrientes (row 25) - Efectivo (row 24) - Efectivo Restringido (row 23)
    formula = f"='1. Balance General'!{col_letter}25-'1. Balance General'!{col_letter}24-'1. Balance General'!{col_letter}23"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '#,##0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
current_row += 1

# Pasivos Corrientes Operativos
ws_ind.cell(current_row, 1, 'Pasivos Corrientes Operativos')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # Total Pasivos Corrientes (row 57) - Operaciones de financiamiento (row 49)
    formula = f"='1. Balance General'!{col_letter}57-'1. Balance General'!{col_letter}49"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '#,##0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
current_row += 1

# KTNO = Activos Corrientes Operativos - Pasivos Corrientes Operativos
ws_ind.cell(current_row, 1, 'KTNO')
ws_ind.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{current_row-2}-{col_letter}{current_row-1}"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '#,##0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
    ws_ind.cell(current_row, col_idx).fill = indicator_fill
    ws_ind.cell(current_row, col_idx).font = bold_font
ktno_row = current_row
current_row += 2

# ΔKTNO (Variation)
ws_ind.cell(current_row, 1, 'ΔKTNO (Variación del KTNO)')
ws_ind.cell(current_row, 1).font = bold_font
ws_ind.cell(current_row, 2, 'N/A')
ws_ind.cell(current_row, 2).alignment = center_align
for col_idx, col_letter in enumerate(['C', 'D', 'E', 'F', 'G'], start=3):
    prev_letter = chr(ord(col_letter)-1)
    formula = f"={col_letter}{ktno_row}-{prev_letter}{ktno_row}"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '#,##0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
    ws_ind.cell(current_row, col_idx).fill = indicator_fill
    ws_ind.cell(current_row, col_idx).font = bold_font
delta_ktno_row = current_row
current_row += 2

# ANDEO - Activos Netos de Operación
ws_ind.cell(current_row, 1, 'ANDEO (Activos Netos de Operación)')
ws_ind.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # Total Activos - Pasivos que no generan intereses (Pasivos Corrientes sin financiamiento)
    # Simplificado: Total Activos - Pasivos Corrientes Operativos
    formula = f"='1. Balance General'!{col_letter}27-{col_letter}{current_row-4}"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '#,##0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
    ws_ind.cell(current_row, col_idx).fill = indicator_fill
    ws_ind.cell(current_row, col_idx).font = bold_font
current_row += 2

# Ciclo de Conversión de Efectivo (CCE)
ws_ind.cell(current_row, 1, 'CICLO DE CONVERSIÓN DE EFECTIVO (CCE)')
ws_ind.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_ind.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_ind.cell(current_row, col).fill = subheader_fill
current_row += 1

# Días de inventario
ws_ind.cell(current_row, 1, 'Días de Inventario')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # (Inventario / Costo de Ventas) * 365
    # Inventario row 16, Costo de Ventas row 7 of ER
    formula = f"=(('1. Balance General'!{col_letter}16/ABS('2. Estado de Resultados'!{col_letter}7))*365)"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '0.0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
dias_inv_row = current_row
current_row += 1

# Días de cartera
ws_ind.cell(current_row, 1, 'Días de Cartera (CxC)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # (Cuentas por Cobrar / Ventas) * 365
    # CxC corrientes row 18, Ventas row 6 of ER
    formula = f"=(('1. Balance General'!{col_letter}18/'2. Estado de Resultados'!{col_letter}6)*365)"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '0.0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
dias_cartera_row = current_row
current_row += 1

# Días de proveedores
ws_ind.cell(current_row, 1, 'Días de Proveedores (CxP)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # (Cuentas por Pagar / Costo de Ventas) * 365
    # CxP corrientes row 50, Costo de Ventas row 7 of ER
    formula = f"=(('1. Balance General'!{col_letter}50/ABS('2. Estado de Resultados'!{col_letter}7))*365)"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '0.0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
dias_prov_row = current_row
current_row += 1

# CCE = Días Inventario + Días Cartera - Días Proveedores
ws_ind.cell(current_row, 1, 'CCE (días)')
ws_ind.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{dias_inv_row}+{col_letter}{dias_cartera_row}-{col_letter}{dias_prov_row}"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '0.0'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
    ws_ind.cell(current_row, col_idx).fill = indicator_fill
    ws_ind.cell(current_row, col_idx).font = bold_font
current_row += 2

# Productividad del Capital de Trabajo (PKT)
ws_ind.cell(current_row, 1, 'PKT - Productividad del Capital de Trabajo')
ws_ind.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # Ventas / KTNO
    formula = f"='2. Estado de Resultados'!{col_letter}6/{col_letter}{ktno_row}"
    ws_ind.cell(current_row, col_idx, formula)
    ws_ind.cell(current_row, col_idx).number_format = '0.00'
    ws_ind.cell(current_row, col_idx).alignment = right_align
    ws_ind.cell(current_row, col_idx).border = thin_border
    ws_ind.cell(current_row, col_idx).fill = indicator_fill
    ws_ind.cell(current_row, col_idx).font = bold_font

# Adjust column widths
ws_ind.column_dimensions['A'].width = 50
for col in range(2, 8):
    ws_ind.column_dimensions[get_column_letter(col)].width = 16

print("   ✅ Indicadores Financieros creados con fórmulas")

# ==================== SHEET 8: FLUJOS DE CAJA ====================
print("📊 Creando hoja 8: Flujos de Caja...")
ws_fc = wb.create_sheet("8. Flujos de Caja")
ws_fc.sheet_view.showGridLines = False

# Title
ws_fc['A1'] = 'ISAGEN S.A. E.S.P.'
ws_fc['A1'].font = Font(name='Arial', size=14, bold=True)
ws_fc['A2'] = 'FLUJOS DE CAJA'
ws_fc['A2'].font = Font(name='Arial', size=12, bold=True)
ws_fc['A3'] = 'Valores en millones de pesos colombianos (COP)'
ws_fc['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_fc.cell(row, 1, 'CONCEPTO')
ws_fc.cell(row, 2, '2020')
ws_fc.cell(row, 3, '2021')
ws_fc.cell(row, 4, '2022')
ws_fc.cell(row, 5, '2023')
ws_fc.cell(row, 6, '2024')
ws_fc.cell(row, 7, '2025 (Sep)')
format_header(ws_fc, row, 7)

current_row = 6

# FCO - Flujo de Caja Operacional
ws_fc.cell(current_row, 1, 'A. FLUJO DE CAJA OPERACIONAL (FCO)')
ws_fc.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_fc.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_fc.cell(current_row, col).fill = subheader_fill
current_row += 1

# EBITDA
ws_fc.cell(current_row, 1, 'EBITDA')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"='7. Indicadores Financieros'!{col_letter}6"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
ebitda_fc_row = current_row
current_row += 1

# (-) Impuestos operacionales
ws_fc.cell(current_row, 1, '(-) Impuesto de Renta Total')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # Impuesto corriente (row 20) + Impuesto diferido (row 21) of ER
    formula = f"='2. Estado de Resultados'!{col_letter}20+'2. Estado de Resultados'!{col_letter}21"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
impuestos_row = current_row
current_row += 1

# (-) ΔKTNO
ws_fc.cell(current_row, 1, '(-) ΔKTNO')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"='7. Indicadores Financieros'!{col_letter}{delta_ktno_row}"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
dktno_fc_row = current_row
current_row += 1

# FCO
ws_fc.cell(current_row, 1, '= FCO (Flujo de Caja Operacional)')
ws_fc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{ebitda_fc_row}+{col_letter}{impuestos_row}-{col_letter}{dktno_fc_row}"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
    ws_fc.cell(current_row, col_idx).fill = cashflow_fill
    ws_fc.cell(current_row, col_idx).font = bold_font
fco_row = current_row
current_row += 2

# FCI - Flujo de Caja de Inversión
ws_fc.cell(current_row, 1, 'B. FLUJO DE CAJA DE INVERSIÓN (FCI)')
ws_fc.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_fc.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_fc.cell(current_row, col).fill = subheader_fill
current_row += 1

# CAPEX (approximated as change in PPE + Depreciation)
ws_fc.cell(current_row, 1, 'CAPEX (Inversión en PPE)')
ws_fc.cell(current_row, 2, 'N/A')
ws_fc.cell(current_row, 2).alignment = center_align
for col_idx, col_letter in enumerate(['C', 'D', 'E', 'F', 'G'], start=3):
    prev_letter = chr(ord(col_letter)-1)
    # CAPEX = (PPE_actual - PPE_anterior) + Depreciación
    formula = f"=('1. Balance General'!{col_letter}6-'1. Balance General'!{prev_letter}6)+'2. Estado de Resultados'!{col_letter}17"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
capex_row = current_row
current_row += 1

# FCI  (generalmente negativo)
ws_fc.cell(current_row, 1, '= FCI (Flujo de Caja de Inversión)')
ws_fc.cell(current_row, 1).font = bold_font
ws_fc.cell(current_row, 2, 'N/A')
ws_fc.cell(current_row, 2).alignment = center_align
for col_idx, col_letter in enumerate(['C', 'D', 'E', 'F', 'G'], start=3):
    formula = f"=-{col_letter}{capex_row}"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
    ws_fc.cell(current_row, col_idx).fill = cashflow_fill
    ws_fc.cell(current_row, col_idx).font = bold_font
fci_row = current_row
current_row += 2

# FCFF - Free Cash Flow to the Firm
ws_fc.cell(current_row, 1, 'C. FCFF - FREE CASH FLOW TO THE FIRM')
ws_fc.cell(current_row, 1).font = bold_font
ws_fc.cell(current_row, 1).fill = total_fill
ws_fc.cell(current_row, 2, 'N/A')
ws_fc.cell(current_row, 2).alignment = center_align
for col_idx, col_letter in enumerate(['C', 'D', 'E', 'F', 'G'], start=3):
    formula = f"={col_letter}{fco_row}+{col_letter}{fci_row}"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
    ws_fc.cell(current_row, col_idx).fill = total_fill
    ws_fc.cell(current_row, col_idx).font = bold_font
fcff_row = current_row
current_row += 2

# FCD - Flujo de Caja de la Deuda
ws_fc.cell(current_row, 1, 'D. FLUJO DE CAJA DE LA DEUDA (FCD)')
ws_fc.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_fc.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_fc.cell(current_row, col).fill = subheader_fill
current_row += 1

# Variación de la deuda
ws_fc.cell(current_row, 1, 'Variación Deuda Financiera')
ws_fc.cell(current_row, 2, 'N/A')
ws_fc.cell(current_row, 2).alignment = center_align
for col_idx, col_letter in enumerate(['C', 'D', 'E', 'F', 'G'], start=3):
    prev_letter = chr(ord(col_letter)-1)
    # Cambio en deuda total (corriente + no corriente)
    # Deuda NC row 38, Deuda C row 49
    formula = f"=(('1. Balance General'!{col_letter}38+'1. Balance General'!{col_letter}49)-('1. Balance General'!{prev_letter}38+'1. Balance General'!{prev_letter}49))"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
var_deuda_row = current_row
current_row += 1

# Intereses pagados (con escudo fiscal)
ws_fc.cell(current_row, 1, '(-) Intereses × (1-T)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # Gastos financieros row 13 * 0.65 (asumiendo tasa impuesto 35%)
    formula = f"='2. Estado de Resultados'!{col_letter}13*0.65"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
intereses_row = current_row
current_row += 1

# FCD
ws_fc.cell(current_row, 1, '= FCD (Flujo de Caja de la Deuda)')
ws_fc.cell(current_row, 1).font = bold_font
ws_fc.cell(current_row, 2, 'N/A')
ws_fc.cell(current_row, 2).alignment = center_align
for col_idx, col_letter in enumerate(['C', 'D', 'E', 'F', 'G'], start=3):
    formula = f"={col_letter}{var_deuda_row}+{col_letter}{intereses_row}"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
    ws_fc.cell(current_row, col_idx).fill = cashflow_fill
    ws_fc.cell(current_row, col_idx).font = bold_font
fcd_row = current_row
current_row += 2

# FCFE - Free Cash Flow to Equity
ws_fc.cell(current_row, 1, 'E. FCFE - FREE CASH FLOW TO EQUITY')
ws_fc.cell(current_row, 1).font = bold_font
ws_fc.cell(current_row, 1).fill = total_fill
ws_fc.cell(current_row, 2, 'N/A')
ws_fc.cell(current_row, 2).alignment = center_align
for col_idx, col_letter in enumerate(['C', 'D', 'E', 'F', 'G'], start=3):
    formula = f"={col_letter}{fcff_row}+{col_letter}{fcd_row}"
    ws_fc.cell(current_row, col_idx, formula)
    ws_fc.cell(current_row, col_idx).number_format = '#,##0'
    ws_fc.cell(current_row, col_idx).alignment = right_align
    ws_fc.cell(current_row, col_idx).border = thin_border
    ws_fc.cell(current_row, col_idx).fill = total_fill
    ws_fc.cell(current_row, col_idx).font = bold_font

# Adjust column widths
ws_fc.column_dimensions['A'].width = 50
for col in range(2, 8):
    ws_fc.column_dimensions[get_column_letter(col)].width = 16

print("   ✅ Flujos de Caja creados con fórmulas")

# Save workbook
wb.save(wb_file)
print(f"\n💾 Archivo actualizado con indicadores y flujos de caja")
print("="*80)
print(f"\n🎉 ANÁLISIS COMPLETO:")
print("   ✅ Hoja 1: Balance General")
print("   ✅ Hoja 2: Estado de Resultados")
print("   ✅ Hoja 3: Análisis Vertical Balance")
print("   ✅ Hoja 4: Análisis Horizontal Balance")
print("   ✅ Hoja 5: Análisis Vertical Estado de Resultados")
print("   ✅ Hoja 6: Análisis Horizontal Estado de Resultados")
print("   ✅ Hoja 7: Indicadores Financieros (EBITDA, KTNO, ΔKTNO, ANDEO, CCE, PKT)")
print("   ✅ Hoja 8: Flujos de Caja (FCO, FCI, FCFF, FCD, FCFE)")
print("\n📂 Archivo: TALLER07_ISAGEN_VALORACION.xlsx")
print("="*80)
