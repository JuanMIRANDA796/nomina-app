import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

print("="*80)
print("AÑADIENDO COSTO DE CAPITAL (WACC) Y VALORACIÓN DCF")
print("="*80)

# Load existing workbook
wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = load_workbook(wb_file)

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
subheader_fill = PatternFill(start_color="7BA0C9", end_color="7BA0C9", fill_type="solid")
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
parameter_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
wacc_fill = PatternFill(start_color="C5E0B4", end_color="C5E0B4", fill_type="solid")
valuation_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
bold_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

# ==================== SHEET 9: COSTO DE CAPITAL Y WACC ====================
print("\n📊 Creando hoja 9: Costo de Capital y WACC...")
ws_wacc = wb.create_sheet("9. Costo de Capital WACC")
ws_wacc.sheet_view.showGridLines = False

# Title
ws_wacc['A1'] = 'ISAGEN S.A. E.S.P.'
ws_wacc['A1'].font = Font(name='Arial', size=14, bold=True)
ws_wacc['A2'] = 'COSTO DE CAPITAL Y WACC'
ws_wacc['A2'].font = Font(name='Arial', size=12, bold=True)
ws_wacc['A3'] = 'Weighted Average Cost of Capital'
ws_wacc['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_wacc.cell(row, 1, 'PARÁMETRO')
ws_wacc.cell(row, 2, '2020')
ws_wacc.cell(row, 3, '2021')
ws_wacc.cell(row, 4, '2022')
ws_wacc.cell(row, 5, '2023')
ws_wacc.cell(row, 6, '2024')
ws_wacc.cell(row, 7, '2025 (Sep)')
format_header(ws_wacc, row, 7)

current_row = 6

# SECTION A: PARÁMETROS DE MERCADO
ws_wacc.cell(current_row, 1, 'A. PARÁMETROS DE MERCADO')
ws_wacc.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_wacc.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_wacc.cell(current_row, col).fill = subheader_fill
current_row += 1

# Rf - Tasa Libre de Riesgo (from sheet 8)
ws_wacc.cell(current_row, 1, 'Rf - Tasa Libre de Riesgo (TES 10 años)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"='8. Flujos de Caja'!{col_letter}28"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = parameter_fill
rf_row = current_row
current_row += 1

# Rm - Retorno del Mercado (from sheet 8)
ws_wacc.cell(current_row, 1, 'Rm - Retorno del Mercado (COLCAP)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"='8. Flujos de Caja'!{col_letter}27"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = parameter_fill
rm_row = current_row
current_row += 1

# Prima de Riesgo de Mercado
ws_wacc.cell(current_row, 1, 'Prima de Riesgo de Mercado (Rm - Rf)')
ws_wacc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{rm_row}-{col_letter}{rf_row}"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = formula_fill
    ws_wacc.cell(current_row, col_idx).font = bold_font
prima_riesgo_row = current_row
current_row += 1

# Riesgo País
ws_wacc.cell(current_row, 1, 'Riesgo País Colombia')
for col_idx in range(2, 8):
    ws_wacc.cell(current_row, col_idx, 0.025)  # 2.5% asumido
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = parameter_fill
riesgo_pais_row = current_row
current_row += 1

# Beta del Sector (from sheet 8 - CELSIA row)
ws_wacc.cell(current_row, 1, 'Beta del Sector (Comparable: CELSIA)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"='8. Flujos de Caja'!{col_letter}26"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.0000'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = parameter_fill
beta_sector_row = current_row
current_row += 2

# SECTION B: ESTRUCTURA DE CAPITAL
ws_wacc.cell(current_row, 1, 'B. ESTRUCTURA DE CAPITAL')
ws_wacc.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_wacc.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_wacc.cell(current_row, col).fill = subheader_fill
current_row += 1

# Deuda (D)
ws_wacc.cell(current_row, 1, 'D - Deuda Financiera Total')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # Deuda NC (row 38) + Deuda C (row 49) from Balance
    formula = f"='1. Balance General'!{col_letter}38+'1. Balance General'!{col_letter}49"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '#,##0'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
d_row = current_row
current_row += 1

# Patrimonio (E)
ws_wacc.cell(current_row, 1, 'E - Patrimonio Total')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # Total Patrimonio (row 36) from Balance
    formula = f"='1. Balance General'!{col_letter}36"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '#,##0'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
e_row = current_row
current_row += 1

# D + E
ws_wacc.cell(current_row, 1, 'D + E (Valor Total)')
ws_wacc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{d_row}+{col_letter}{e_row}"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '#,##0'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = total_fill
    ws_wacc.cell(current_row, col_idx).font = bold_font
de_row = current_row
current_row += 1

# D/E Ratio
ws_wacc.cell(current_row, 1, 'D/E - Ratio Deuda/Patrimonio')
ws_wacc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{d_row}/{col_letter}{e_row}"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = formula_fill
    ws_wacc.cell(current_row, col_idx).font = bold_font
de_ratio_row = current_row
current_row += 1

# % Deuda
ws_wacc.cell(current_row, 1, '% Deuda = D/(D+E)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{d_row}/{col_letter}{de_row}"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = formula_fill
pct_d_row = current_row
current_row += 1

# % Patrimonio
ws_wacc.cell(current_row, 1, '% Patrimonio = E/(D+E)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{e_row}/{col_letter}{de_row}"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = formula_fill
pct_e_row = current_row
current_row += 2

# SECTION C: COSTO DE LA DEUDA (Kd)
ws_wacc.cell(current_row, 1, 'C. COSTO DE LA DEUDA (Kd)')
ws_wacc.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_wacc.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_wacc.cell(current_row, col).fill = subheader_fill
current_row += 1

# Gastos financieros
ws_wacc.cell(current_row, 1, 'Gastos Financieros')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"='2. Estado de Resultados'!{col_letter}13"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '#,##0'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
gastos_fin_row = current_row
current_row += 1

# Kd (antes de impuestos)
ws_wacc.cell(current_row, 1, 'Kd - Costo de la Deuda (antes de impuestos)')
ws_wacc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"=ABS({col_letter}{gastos_fin_row}/{col_letter}{d_row})"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = formula_fill
    ws_wacc.cell(current_row, col_idx).font = bold_font
kd_row = current_row
current_row += 1

# Tasa de impuestos (T)
ws_wacc.cell(current_row, 1, 'T - Tasa de Impuestos (efectiva)')
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    # (Impuesto corriente + diferido) / Utilidad antes de impuestos
    formula = f"=ABS(('2. Estado de Resultados'!{col_letter}20+'2. Estado de Resultados'!{col_letter}21)/'2. Estado de Resultados'!{col_letter}15)"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = formula_fill
t_row = current_row
current_row += 1

# Kd neto (después de impuestos)
ws_wacc.cell(current_row, 1, 'Kd (neto) = Kd × (1-T)')
ws_wacc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{kd_row}*(1-{col_letter}{t_row})"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = wacc_fill
    ws_wacc.cell(current_row, col_idx).font = bold_font
kd_neto_row = current_row
current_row += 2

# SECTION D: BETA Y COSTO DEL PATRIMONIO (Ke)
ws_wacc.cell(current_row, 1, 'D. BETA Y COSTO DEL PATRIMONIO (Ke)')
ws_wacc.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_wacc.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_wacc.cell(current_row, col).fill = subheader_fill
current_row += 1

# Beta Desapalancado (Bu) - assuming sector beta is levered, we unlever it first
# Bu = BL / [1 + (1-T) × (D/E)]
ws_wacc.cell(current_row, 1, 'Bu - Beta Desapalancado')
ws_wacc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{beta_sector_row}/(1+(1-{col_letter}{t_row})*{col_letter}{de_ratio_row})"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.0000'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = formula_fill
    ws_wacc.cell(current_row, col_idx).font = bold_font
bu_row = current_row
current_row += 1

# Beta Apalancado de ISAGEN (BL) - re-lever with ISAGEN's D/E
# BL = Bu × [1 + (1-T) × (D/E)]
ws_wacc.cell(current_row, 1, 'BL - Beta Apalancado (ISAGEN)')
ws_wacc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{bu_row}*(1+(1-{col_letter}{t_row})*{col_letter}{de_ratio_row})"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.0000'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = formula_fill
    ws_wacc.cell(current_row, col_idx).font = bold_font
bl_row = current_row
current_row += 1

# Ke - Costo del Patrimonio (CAPM)
# Ke = Rf + BL × (Rm - Rf) + Riesgo País
ws_wacc.cell(current_row, 1, 'Ke - Costo del Patrimonio (CAPM)')
ws_wacc.cell(current_row, 1).font = bold_font
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{rf_row}+{col_letter}{bl_row}*{col_letter}{prima_riesgo_row}+{col_letter}{riesgo_pais_row}"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = thin_border
    ws_wacc.cell(current_row, col_idx).fill = wacc_fill
    ws_wacc.cell(current_row, col_idx).font = bold_font
ke_row = current_row
current_row += 2

# SECTION E: WACC
ws_wacc.cell(current_row, 1, 'E. WACC - COSTO PROMEDIO PONDERADO DE CAPITAL')
ws_wacc.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_wacc.cell(current_row, 1).fill = subheader_fill
for col in range(2, 8):
    ws_wacc.cell(current_row, col).fill = subheader_fill
current_row += 1

# WACC = Ke × (E/(D+E)) + Kd(neto) × (D/(D+E))
ws_wacc.cell(current_row, 1, 'WACC = Ke×%E + Kd(neto)×%D')
ws_wacc.cell(current_row, 1).font = Font(name='Arial', size=11, bold=True)
for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
    formula = f"={col_letter}{ke_row}*{col_letter}{pct_e_row}+{col_letter}{kd_neto_row}*{col_letter}{pct_d_row}"
    ws_wacc.cell(current_row, col_idx, formula)
    ws_wacc.cell(current_row, col_idx).number_format = '0.00%'
    ws_wacc.cell(current_row, col_idx).alignment = right_align
    ws_wacc.cell(current_row, col_idx).border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    ws_wacc.cell(current_row, col_idx).fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    ws_wacc.cell(current_row, col_idx).font = Font(name='Arial', size=11, bold=True)
wacc_row = current_row

# Adjust column widths
ws_wacc.column_dimensions['A'].width = 50
for col in range(2, 8):
    ws_wacc.column_dimensions[get_column_letter(col)].width = 16

print("   ✅ Costo de Capital y WACC creado con fórmulas")

# Save workbook
wb.save(wb_file)
print(f"\n💾 Archivo actualizado con WACC")
print("="*80)
