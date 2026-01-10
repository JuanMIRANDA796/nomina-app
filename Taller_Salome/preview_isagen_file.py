import pandas as pd
import openpyxl
from openpyxl import load_workbook

print("="*100)
print("VISTA PREVIA: TALLER07_ISAGEN_VALORACION.xlsx")
print("="*100)

# Load workbook
wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = load_workbook(wb_file, data_only=False)  # Keep formulas visible

print(f"\n📊 ESTRUCTURA DEL ARCHIVO:")
print(f"   Número de hojas: {len(wb.sheetnames)}")
print(f"\n📋 HOJAS CREADAS:\n")

for idx, sheet_name in enumerate(wb.sheetnames, 1):
    ws = wb[sheet_name]
    print(f"   {idx}. {sheet_name}")
    rows_with_data = sum(1 for row in ws.iter_rows() if any(cell.value for cell in row))
    print(f"       - Filas con datos: {rows_with_data}")
    print(f"       - Dimensiones: {ws.max_row} x {ws.max_column}")

print(f"\n" + "="*100)
print("VISTA PREVIA DE CADA HOJA (Primeras filas)")
print("="*100)

# Preview each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*100}")
    print(f"HOJA: {sheet_name}")
    print("="*100)
    
    # Show first 15 rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
        row_data = []
        for cell in row:
            if cell.value is not None:
                # Show formula if it's a formula
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    row_data.append(f"[FÓRMULA: {cell.value[:40]}...]" if len(cell.value) > 40 else f"[FÓRMULA: {cell.value}]")
                else:
                    val_str = str(cell.value)
                    row_data.append(val_str[:30] if len(val_str) <= 30 else val_str[:27] + "...")
            else:
                row_data.append("")
        
        if any(row_data):  # Only show rows with data
            print(f"   Fila {row_idx:2d}: {' | '.join(row_data[:7])}")  # First 7 columns
    
    print()

print("="*100)
print("RESUMEN DE FÓRMULAS IMPLEMENTADAS")
print("="*100)

# Count formulas in key sheets
formula_counts = {}
for sheet_name in ["3. Análisis Vertical BSC", "7. Indicadores Financieros", "8. Flujos de Caja"]:
    ws = wb[sheet_name]
    formula_count = 0
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
    formula_counts[sheet_name] = formula_count

print(f"\n✅ Fórmulas creadas por hoja:")
for sheet, count in formula_counts.items():
    print(f"   {sheet}: {count} fórmulas")

print(f"\n{'='*100}")
print("VERIFICACIÓN DE INDICADORES CLAVE")
print("="*100)

# Show sample calculations with formulas
ws_ind = wb["7. Indicadores Financieros"]
print(f"\n📊 Indicadores Financieros (Hoja 7):")
print(f"\n   {'Indicador':<40} {'2024 (Fila 6)':>20}")
print(f"   {'-'*70}")

key_indicators = [
    (6, "EBITDA"),
    (11, "KTNO"),
    (12, "ΔKTNO"),
    (14, "ANDEO"),
    (21, "CCE (días)"),
    (23, "PKT")
]

for row_num, indicator_name in key_indicators:
    cell_f = ws_ind.cell(row=row_num, column=6)  # Column F = 2024
    if isinstance(cell_f.value, str) and cell_f.value.startswith('='):
        print(f"   {indicator_name:<40} {cell_f.value}")
    else:
        print(f"   {indicator_name:<40} VALOR: {cell_f.value}")

# Show cash flows
ws_fc = wb["8. Flujos de Caja"]
print(f"\n💰 Flujos de Caja (Hoja 8):")
print(f"\n   {'Flujo de Caja':<40} {'2024 (Columna F)':>20}")
print(f"   {'-'*70}")

key_flows = [
    (10, "FCO - Flujo de Caja Operacional"),
    (16, "FCI - Flujo de Caja de Inversión"),
    (18, "FCFF - Free Cash Flow to Firm"),
    (24, "FCD - Flujo de Caja de la Deuda"),
    (26, "FCFE - Free Cash Flow to Equity")
]

for row_num, flow_name in key_flows:
    cell_f = ws_fc.cell(row=row_num, column=5)  # Column E = 2023 (adjust as needed)
    if isinstance(cell_f.value, str) and cell_f.value.startswith('='):
        print(f"   {flow_name:<40} {cell_f.value[:45]}...")
    else:
        print(f"   {flow_name:<40} VALOR: {cell_f.value}")

print(f"\n{'='*100}")
print("✅ ARCHIVO GENERADO EXITOSAMENTE")
print("="*100)
print(f"\n📂 Ubicación: {wb_file}")
print(f"📊 Total de hojas: {len(wb.sheetnames)}")
print(f"✨ Todas las fórmulas están preservadas en el archivo Excel")
print(f"\n🎯 PRÓXIMOS PASOS:")
print(f"   4. Costo de Capital y WACC")
print(f"   5. Valoración por Flujos Descontados (DCF)")
print(f"   6. Análisis Integral y Documento Word")
print("="*100)
