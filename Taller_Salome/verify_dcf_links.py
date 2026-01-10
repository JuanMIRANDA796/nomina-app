import openpyxl
import pandas as pd

file_path = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True) # Data only to see values

print("--- SHEET: 10. Valoración DCF (Verified) ---")
if '10. Valoración DCF' in wb.sheetnames:
    ws = wb['10. Valoración DCF']
    
    # Read FCFF section (around row 25 where we inserted)
    # let's look for "FCFF"
    start_row = 1
    for r in range(1, 40):
        if ws.cell(r,1).value and "FCFF" in str(ws.cell(r,1).value) and "CRECIMIENTO" not in str(ws.cell(r,1).value).upper():
            start_row = r
            break
            
    print(f"FCFF Section starts approx row {start_row}")
    
    # Print 2026 values (Column E)
    print("\nValues for 2026 (Projected):")
    rows_to_check = [start_row+1, start_row+2, start_row+3, start_row+4, start_row+5]
    labels = ["EBITDA", "Tax", "Delta KTNO", "Capex", "Total FCFF"]
    
    for r, label in zip(rows_to_check, labels):
        val = ws.cell(r, 5).value # Col E
        print(f"{label}: {val}")

    # Print Enterprise Value
    print("\nEnterprise Value & Equity Value:")
    # Look for "VALOR DE OPERACIÓN"
    for r in range(start_row, 50):
        val = ws.cell(r,1).value
        val_next = ws.cell(r+3, 1).value # usually equity is a few rows down
        if val and "OPERACIÓN" in str(val).upper():
             print(f"Enterprise Value: {ws.cell(r, 14).value}") # Col N usually
        if val and "PATRIMONIAL" in str(val).upper():
             print(f"Equity Value: {ws.cell(r, 14).value}")

else:
    print("Sheet 10 not found")
