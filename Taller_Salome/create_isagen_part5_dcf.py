import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("="*80)
print("AÑADIENDO VALORACIÓN POR FLUJOS DESCONTADOS (DCF)")
print("="*80)

# Load existing workbook
wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = load_workbook(wb_file)

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
subheader_fill = PatternFill(start_color="7BA0C9", end_color="7BA0C9", fill_type="solid")
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
projection_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
terminal_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
valuation_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
bold_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
thick_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

def format_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

# ==================== SHEET 10: PROYECCIONES Y VALORACIÓN DCF ====================
print("\n📊 Creando hoja 10: Proyecciones y Valoración DCF...")
ws_dcf = wb.create_sheet("10. Valoración DCF")
ws_dcf.sheet_view.showGridLines = False

# Title
ws_dcf['A1'] = 'ISAGEN S.A. E.S.P.'
ws_dcf['A1'].font = Font(name='Arial', size=14, bold=True)
ws_dcf['A2'] = 'VALORACIÓN POR FLUJOS DE CAJA DESCONTADOS (DCF)'
ws_dcf['A2'].font = Font(name='Arial', size=12, bold=True)
ws_dcf['A3'] = 'Proyecciones 2026-2035 y Valor Terminal'
ws_dcf['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers - Include historical years and projections
row = 5
ws_dcf.cell(row, 1, 'CONCEPTO')
ws_dcf.cell(row, 2, '2023')
ws_dcf.cell(row, 3, '2024')
ws_dcf.cell(row, 4, '2025')
ws_dcf.cell(row, 5, '2026')
ws_dcf.cell(row, 6, '2027')
ws_dcf.cell(row, 7, '2028')
ws_dcf.cell(row, 8, '2029')
ws_dcf.cell(row, 9, '2030')
ws_dcf.cell(row, 10, '2031')
ws_dcf.cell(row, 11, '2032')
ws_dcf.cell(row, 12, '2033')
ws_dcf.cell(row, 13, '2034')
ws_dcf.cell(row, 14, '2035')
format_header(ws_dcf, row, 14)

current_row = 6

# SECTION A: SUPUESTOS DE PROYECCIÓN
ws_dcf.cell(current_row, 1, 'A. SUPUESTOS DE PROYECCIÓN')
ws_dcf.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_dcf.cell(current_row, 1).fill = subheader_fill
for col in range(2, 15):
    ws_dcf.cell(current_row, col).fill = subheader_fill
current_row += 1

# Tasa de crecimiento de FCFF
ws_dcf.cell(current_row, 1, 'Tasa de Crecimiento FCFF (g)')
# Historical - calculate based on actual growth
for col_idx in range(2, 5):  # 2023, 2024, 2025
    ws_dcf.cell(current_row, col_idx, "N/A")
    ws_dcf.cell(current_row, col_idx).alignment = center_align
# Projections - assume 4% growth
for col_idx in range(5, 15):  # 2026-2035
    ws_dcf.cell(current_row, col_idx, 0.04)
    ws_dcf.cell(current_row, col_idx).number_format = '0.00%'
    ws_dcf.cell(current_row, col_idx).alignment = right_align
    ws_dcf.cell(current_row, col_idx).border = thin_border
    ws_dcf.cell(current_row, col_idx).fill = projection_fill
growth_rate_row = current_row
current_row += 1

# Tasa de crecimiento perpetuo (g terminal)
ws_dcf.cell(current_row, 1, 'Tasa de Crecimiento Perpetuo (g terminal)')
for col_idx in range(2, 15):
    ws_dcf.cell(current_row, col_idx, 0.035)  # 3.5%
    ws_dcf.cell(current_row, col_idx).number_format = '0.00%'
    ws_dcf.cell(current_row, col_idx).alignment = right_align
    ws_dcf.cell(current_row, col_idx).border = thin_border
    ws_dcf.cell(current_row, col_idx).fill = projection_fill
g_terminal_row = current_row
current_row += 2

# SECTION B: FCFF PROYECTADO
ws_dcf.cell(current_row, 1, 'B. FCFF - FREE CASH FLOW TO THE FIRM')
ws_dcf.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_dcf.cell(current_row, 1).fill = subheader_fill
for col in range(2, 15):
    ws_dcf.cell(current_row, col).fill = subheader_fill
current_row += 1

# FCFF Historical and Projected
ws_dcf.cell(current_row, 1, 'FCFF')
ws_dcf.cell(current_row, 1).font = bold_font
# Historical from sheet 8
ws_dcf.cell(current_row, 2, "='8. Flujos de Caja'!E18")  # 2023
ws_dcf.cell(current_row, 3, "='8. Flujos de Caja'!F18")  # 2024
ws_dcf.cell(current_row, 4, "='8. Flujos de Caja'!G18")  # 2025

# Projections - grow at assumed rate
for col_idx in range(5, 15):  # 2026-2035
    prev_col_letter = get_column_letter(col_idx - 1)
    col_letter = get_column_letter(col_idx)
    formula = f"={prev_col_letter}{current_row}*(1+{col_letter}{growth_rate_row})"
    ws_dcf.cell(current_row, col_idx, formula)

# Format all FCFF cells
for col_idx in range(2, 15):
    ws_dcf.cell(current_row, col_idx).number_format = '#,##0'
    ws_dcf.cell(current_row, col_idx).alignment = right_align
    ws_dcf.cell(current_row, col_idx).border = thin_border
    ws_dcf.cell(current_row, col_idx).fill = formula_fill if col_idx >= 5 else total_fill
    ws_dcf.cell(current_row, col_idx).font = bold_font
fcff_row = current_row
current_row += 2

# SECTION C: WACC Y DESCUENTO
ws_dcf.cell(current_row, 1, 'C. TASA DE DESCUENTO (WACC)')
ws_dcf.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_dcf.cell(current_row, 1).fill = subheader_fill
for col in range(2, 15):
    ws_dcf.cell(current_row, col).fill = subheader_fill
current_row += 1

# WACC
ws_dcf.cell(current_row, 1, 'WACC')
ws_dcf.cell(current_row, 1).font = bold_font
# Historical from sheet 9
ws_dcf.cell(current_row, 2, "='9. Costo de Capital WACC'!E39")  # 2023
ws_dcf.cell(current_row, 3, "='9. Costo de Capital WACC'!F39")  # 2024
ws_dcf.cell(current_row, 4, "='9. Costo de Capital WACC'!G39")  # 2025

# Projections - use  2024 WACC as baseline
for col_idx in range(5, 15):  # 2026-2035
    ws_dcf.cell(current_row, col_idx, "='9. Costo de Capital WACC'!F39")  # Use 2024 WACC

# Format all WACC cells
for col_idx in range(2, 15):
    ws_dcf.cell(current_row, col_idx).number_format = '0.00%'
    ws_dcf.cell(current_row, col_idx).alignment = right_align
    ws_dcf.cell(current_row, col_idx).border = thin_border
    ws_dcf.cell(current_row, col_idx).fill = formula_fill
    ws_dcf.cell(current_row, col_idx).font = bold_font
wacc_dcf_row = current_row
current_row += 1

# Período
ws_dcf.cell(current_row, 1, 'Período (años desde 2025)')
for col_idx, year in enumerate(range(2023, 2036), start=2):
    period = col_idx - 4  # 2025 is period 0
    ws_dcf.cell(current_row, col_idx, period)
    ws_dcf.cell(current_row, col_idx).number_format = '0'
    ws_dcf.cell(current_row, col_idx).alignment = right_align
    ws_dcf.cell(current_row, col_idx).border = thin_border
period_row = current_row
current_row += 1

# Factor de descuento = 1 / (1 + WACC)^período
ws_dcf.cell(current_row, 1, 'Factor de Descuento = 1/(1+WACC)^n')
for col_idx in range(2, 15):
    col_letter = get_column_letter(col_idx)
    formula = f"=1/((1+{col_letter}{wacc_dcf_row})^{col_letter}{period_row})"
    ws_dcf.cell(current_row, col_idx, formula)
    ws_dcf.cell(current_row, col_idx).number_format = '0.000000'
    ws_dcf.cell(current_row, col_idx).alignment = right_align
    ws_dcf.cell(current_row, col_idx).border = thin_border
    ws_dcf.cell(current_row, col_idx).fill = formula_fill
factor_desc_row = current_row
current_row += 1

# FCFF Descontado (Valor Presente)
ws_dcf.cell(current_row, 1, 'FCFF Descontado (VP)')
ws_dcf.cell(current_row, 1).font = bold_font
for col_idx in range(2, 15):
    col_letter = get_column_letter(col_idx)
    formula = f"={col_letter}{fcff_row}*{col_letter}{factor_desc_row}"
    ws_dcf.cell(current_row, col_idx, formula)
    ws_dcf.cell(current_row, col_idx).number_format = '#,##0'
    ws_dcf.cell(current_row, col_idx).alignment = right_align
    ws_dcf.cell(current_row, col_idx).border = thin_border
    ws_dcf.cell(current_row, col_idx).fill = formula_fill
    ws_dcf.cell(current_row, col_idx).font = bold_font
fcff_vp_row = current_row
current_row += 2

# SECTION D: VALOR TERMINAL
ws_dcf.cell(current_row, 1, 'D. VALOR TERMINAL (2035)')
ws_dcf.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_dcf.cell(current_row, 1).fill = subheader_fill
for col in range(2, 15):
    ws_dcf.cell(current_row, col).fill = subheader_fill
current_row += 1

# FCFF 2036 (primer año después del horizonte)
ws_dcf.cell(current_row, 1, 'FCFF 2036 = FCFF2035 × (1+g)')
last_col_letter = get_column_letter(14)  # Column N (2035)
ws_dcf.cell(current_row, 14, f"={last_col_letter}{fcff_row}*(1+{last_col_letter}{g_terminal_row})")
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thin_border
ws_dcf.cell(current_row, 14).fill = terminal_fill
ws_dcf.cell(current_row, 14).font = bold_font
fcff_2036_row = current_row
current_row += 1

# Valor Terminal = FCFF2036 / (WACC - g)
ws_dcf.cell(current_row, 1, 'Valor Terminal = FCFF2036 / (WACC - g)')
ws_dcf.cell(current_row, 1).font = bold_font
ws_dcf.cell(current_row, 14, f"=N{fcff_2036_row}/(N{wacc_dcf_row}-N{g_terminal_row})")
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thick_border
ws_dcf.cell(current_row, 14).fill = terminal_fill
ws_dcf.cell(current_row, 14).font = Font(name='Arial', size=11, bold=True)
tv_row = current_row
current_row += 1

# Valor Terminal Descontado
ws_dcf.cell(current_row, 1, 'Valor Terminal Descontado (VP)')
ws_dcf.cell(current_row, 1).font = bold_font
ws_dcf.cell(current_row, 14, f"=N{tv_row}*N{factor_desc_row}")
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thick_border
ws_dcf.cell(current_row, 14).fill = terminal_fill
ws_dcf.cell(current_row, 14).font = Font(name='Arial', size=11, bold=True)
tv_vp_row = current_row
current_row += 2

# SECTION E: VALOR DE LA EMPRESA
ws_dcf.cell(current_row, 1, 'E. VALORACIÓN DE LA EMPRESA')
ws_dcf.cell(current_row, 1).font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
ws_dcf.cell(current_row, 1).fill = subheader_fill
for col in range(2, 15):
    ws_dcf.cell(current_row, col).fill = subheader_fill
current_row += 1

# Suma de VP de FCFF proyectados (2026-2035)
ws_dcf.cell(current_row, 1, 'Valor Presente FCFF Proyectados')
ws_dcf.cell(current_row, 14, f"=SUM(E{fcff_vp_row}:N{fcff_vp_row})")
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thin_border
ws_dcf.cell(current_row, 14).fill = valuation_fill
ws_dcf.cell(current_row, 14).font = bold_font
vp_fcff_sum_row = current_row
current_row += 1

# (+) Valor Terminal Descontado
ws_dcf.cell(current_row, 1, '(+) Valor Terminal Descontado')
ws_dcf.cell(current_row, 14, f"=N{tv_vp_row}")
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thin_border
ws_dcf.cell(current_row, 14).fill = valuation_fill
ws_dcf.cell(current_row, 14).font = bold_font
current_row += 1

# = VALOR DE OPERACIÓN (Enterprise Value)
ws_dcf.cell(current_row, 1, '= VALOR DE OPERACIÓN (Enterprise Value)')
ws_dcf.cell(current_row, 1).font = Font(name='Arial', size=11, bold=True)
ws_dcf.cell(current_row, 14, f"=N{vp_fcff_sum_row}+N{tv_vp_row}")
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thick_border
ws_dcf.cell(current_row, 14).fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
ws_dcf.cell(current_row, 14).font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
enterprise_value_row = current_row
current_row += 2

# (-) Deuda Financiera Neta
ws_dcf.cell(current_row, 1, '(-) Deuda Financiera Neta')
ws_dcf.cell(current_row, 14, "='1. Balance General'!F38+'1. Balance General'!F49-'1. Balance General'!F24")
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thin_border
deuda_neta_row = current_row
current_row += 1

# (+) Activos No Operacionales
ws_dcf.cell(current_row, 1, '(+) Activos No Operacionales')
ws_dcf.cell(current_row, 14, 0)  # Assuming none for now
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thin_border
activos_no_op_row = current_row
current_row += 1

# = VALOR PATRIMONIAL (Equity Value)
ws_dcf.cell(current_row, 1, '= VALOR PATRIMONIAL (Equity Value)')
ws_dcf.cell(current_row, 1).font = Font(name='Arial', size=11, bold=True)
ws_dcf.cell(current_row, 14, f"=N{enterprise_value_row}-N{deuda_neta_row}+N{activos_no_op_row}")
ws_dcf.cell(current_row, 14).number_format = '#,##0'
ws_dcf.cell(current_row, 14).alignment = right_align
ws_dcf.cell(current_row, 14).border = thick_border
ws_dcf.cell(current_row, 14).fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
ws_dcf.cell(current_row, 14).font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
equity_value_row = current_row

# Adjust column widths
ws_dcf.column_dimensions['A'].width = 50
for col in range(2, 15):
    ws_dcf.column_dimensions[get_column_letter(col)].width = 14

print("   ✅ Valoración DCF creada con fórmulas")

# Save workbook
wb.save(wb_file)
print(f"\n💾 Archivo actualizado con Valoración DCF")
print("="*80)
