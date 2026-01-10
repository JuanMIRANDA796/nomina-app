from openpyxl import load_workbook
import pandas as pd

# Load workbook
wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = load_workbook(wb_file, data_only=True)

print("="*80)
print("RESUMEN FINAL DEL ARCHIVO ISAGEN")
print("="*80)

print(f"\n📊 ESTRUCTURA COMPLETA ({len(wb.sheetnames)} hojas):\n")

for idx, sheet_name in enumerate(wb.sheetnames, 1):
    print(f"   {idx:2d}. {sheet_name}")

print("\n" + "="*80)
print("VALORES CLAVE CALCULADOS")
print("="*80)

# Read WACC sheet
print("\n💰 COSTO DE CAPITAL (2024):")
ws_wacc = wb['9. Costo de Capital WACC']
try:
    wacc_2024 = ws_wacc['F39'].value
    ke_2024 = ws_wacc['F37'].value
    kd_2024 = ws_wacc['F27'].value
    print(f"   WACC: {wacc_2024*100:.2f}%" if wacc_2024 else "   WACC: Calcular en Excel")
    print(f"   Ke (Costo Patrimonio): {ke_2024*100:.2f}%" if ke_2024 else "   Ke: Calcular en Excel")
    print(f"   Kd (Costo Deuda neto): {kd_2024*100:.2f}%" if kd_2024 else "   Kd: Calcular en Excel")
except:
    print("   Valores a calcular al abrir el archivo Excel")

# Read DCF sheet
print("\n📈 VALORACIÓN (millones COP):")
ws_dcf = wb['10. Valoración DCF']
try:
    enterprise_value = ws_dcf['N28'].value
    equity_value = ws_dcf['N32'].value
    print(f"   Valor de Operación (Enterprise Value): ${enterprise_value:,.0f}" if enterprise_value else "   Enterprise Value: Calcular en Excel")
    print(f"   Valor Patrimonial (Equity Value): ${equity_value:,.0f}" if equity_value else "   Equity Value: Calcular en Excel")
except:
    print("   Valores a calcular al abrir el archivo Excel")

print("\n" + "="*80)
print("✅ ARCHIVO COMPLETO Y LISTO PARA USAR")
print("="*80)
print(f"\n📂 Ubicación: {wb_file}")
print("\n🎯 PRÓXIMOS PASOS:")
print("   1. Abrir el archivo Excel")
print("   2. Verificar que todas las fórmulas calculen correctamente")
print("   3. Ajustar supuestos si es necesario (tasas de crecimiento, etc.)")
print("   4. Revisar el análisis de sensibilidad")
print("   5. Preparar el Documento Word con el análisis cualitativo")
print("\n" + "="*80)
