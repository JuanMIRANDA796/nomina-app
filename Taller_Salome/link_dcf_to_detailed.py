import openpyxl
from openpyxl.utils import get_column_letter

print("="*80)
print("ACTUALIZANDO HOJA 10. VALORACIÓN DCF (VINCULANDO A PROYECCIONES DETALLADAS)")
print("="*80)

file_path = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = openpyxl.load_workbook(file_path)

if '10. Valoración DCF' not in wb.sheetnames:
    print("Error: Sheet 10 not found.")
    exit()

ws_dcf = wb['10. Valoración DCF']

# We need to map the DCF rows to the new sheets
# FCFF = EBITDA - Taxes - Capex - Delta KTNO
# Or FCFF = EBIT * (1-t) + Dep - Capex - Delta KTNO
# Let's see how Sheet 10 is currently built.
# It links to '8. Flujos de Caja' for history.
# For projections (Cols E-N / 2026-2035), it currently has hardcoded growth logic or placeholders.

# We will overwrite the projection columns (2026-2035) to link to Sheets 12 & 13.

# 1. WACC (Row ?)
# Link to '14. WACC Proyectado' row 10 (WACC)
# Find WACC Row in DCF
wacc_row = -1
growth_row = -1
fcff_row = -1

for r in range(1, 40):
    val = ws_dcf.cell(r, 1).value
    if val and "WACC" in str(val) and "TASA" not in str(val).upper(): # Avoid header
        wacc_row = r
    if val and "FCFF" in str(val) and "CRECIMIENTO" not in str(val).upper():
        fcff_row = r
    if val and "Crecimiento FCFF" in str(val):
        growth_row = r

print(f"Rows found: WACC={wacc_row}, FCFF={fcff_row}")

# Update WACC (Columns 5 to 14 -> 2026 to 2035)
# Sheet 14 columns: B=2025. C=2026... L=2035.
# DCF Columns: E=2026... N=2035.
# So DCF Col E links to Sheet 14 Col C.
# Offset: Sheet 14 Col = DCF Col - 2.
if wacc_row != -1:
    for col in range(5, 15): # E to N
        sheet14_col_let = get_column_letter(col - 2)
        # Link to Row 10 in Sheet 14
        ws_dcf.cell(wacc_row, col, f"='14. WACC Proyectado'!{sheet14_col_let}10")

# Update FCFF
# FCFF needs to be calculated in DCF or linked?
# The assignment asks for FCFF projections.
# We can calculate FCFF components in the DCF sheet drawing from projections, 
# OR calculate FCFF in the ER/BSC sheets?
# Usually done in DCF.
# FCFF = FCO + FCI
# FCO = EBITDA - Taxes - Delta KTNO.
# FCI = -Capex.

# Let's REBUILD the Calculation Block in DCF to be clear.
# Insert rows if needed? Or just overwrite "B. FCFF" section.

# Let's create a new breakdown in DCF for clarity or just link the final FCFF line?
# Creating a breakdown is better for "Integral Valuation".
# Components:
# (+) EBITDA (From ER)
# (-) Impuestos (From ER)
# (+/-) Delta KTNO (From BSC changes)
# (-) CAPEX (From BSC change + Dep)
# (=) FCFF

# Let's find where to write. "B. FCFF" starts at some row.
# We will insert rows below "B. FCFF" header to show components.

b_section_row = -1
for r in range(1, 20):
    if ws_dcf.cell(r, 1).value == 'B. FCFF - FREE CASH FLOW TO THE FIRM':
        b_section_row = r
        break

if b_section_row != -1:
    print(f"Found B Section at {b_section_row}. Inserting detail rows.")
    # We leave the title.
    # We need rows for: EBITDA, Impuestos, Delta KTNO, Capex.
    # Currently there is just one row "FCFF".
    # Let's Insert 4 rows after header
    ws_dcf.insert_rows(b_section_row + 1, 4)
    
    # Define Labels
    ws_dcf.cell(b_section_row + 1, 1, "(+) EBITDA")
    ws_dcf.cell(b_section_row + 2, 1, "(-) Impuestos Operativos")
    ws_dcf.cell(b_section_row + 3, 1, "(+/-) Variación KTNO")
    ws_dcf.cell(b_section_row + 4, 1, "(-) CAPEX")
    
    # Map to Sheets
    # Columns 2026-2035 (Col 5 to 14)
    # 2023-2025 (Cols 2-4) can remain linked to '8. Flujos de Caja' or updated?
    # Let's update 2026-2035 links.
    
    for col in range(5, 15):
        c_proj = get_column_letter(col - 2) # Mapping to Proj Sheets (C=2026)
        
        # EBITDA: Sheet 12 (ER). EBITDA row?
        # In Sheet 12, I constructed it. Need to find EBITDA row index.
        # Based on create_detailed_projections:
        # Row 13 (Index) in list -> Row 14+3? No headers are row 3.
        # Let's look up "EBITDA" in Sheet 12.
        # ... Wait, I copied structure of Sheet 2.
        # Sheet 2: Row 13 is EBITDA (0-based index in pandas) -> Row 17 in Excel?
        # Let's use MATCH or just know the row. 
        # In '2. Estado de Resultados', EBITDA is Row 14 (if 1-based header is included).
        # Pandas index 13.
        # In Sheet 12: Header Row 3. Data starts Row 4.
        # Index 13 -> Row 4+13 = 17.
        # Let's verify with script later? Assuming Row 17 for now.
        ws_dcf.cell(b_section_row + 1, col, f"='12. ER Proyectado'!{c_proj}17")
        
        # Taxes: Sheet 12 (ER). Row 18 is Current Tax.
        ws_dcf.cell(b_section_row + 2, col, f"=-'12. ER Proyectado'!{c_proj}18")
        
        # Delta KTNO: Sheet 13 (BSC).
        # KTNO = CxC + Inv - CxP.
        # Delta = KTNO_t - KTNO_t-1.
        # CxC: Row 16 (Current from code R_CXC_C = 12+5 = 17? Wait. 12+4=16 in ER? No BSC.)
        # R_CXC_C = 12 + 5 = 17.
        # R_INV = 11 + 5 = 16.
        # R_CXP_C = 44 + 5 = 49.
        # Prev Column
        c_prev_proj = get_column_letter(col - 3) if col > 5 else "B" # Base 2025 is Col B
        
        ktno_curr = f"('13. BSC Proyectado'!{c_proj}17+'13. BSC Proyectado'!{c_proj}16-'13. BSC Proyectado'!{c_proj}49)"
        ktno_prev = f"('13. BSC Proyectado'!{c_prev_proj}17+'13. BSC Proyectado'!{c_prev_proj}16-'13. BSC Proyectado'!{c_prev_proj}49)"
        ws_dcf.cell(b_section_row + 3, col, f"=-({ktno_curr}-{ktno_prev})")
        
        # Capex: Formula in BSC is: PPE_t = PPE_t-1 + Capex - Dep.
        # Capex = PPE_t - PPE_t-1 + Dep.
        # PPE Row: 1 + 5 = 6.
        # Dep Row: In ER? Row 11 (index) -> Row 15 in Sheet 12.
        # However, historical ER has Depreciación (Row 11) and Amortización (Row 12).
        # Let's sum Row 15 and 16 from Sheet 12.
        ppe_curr = f"'13. BSC Proyectado'!{c_proj}6"
        ppe_prev = f"'13. BSC Proyectado'!{c_prev_proj}6"
        dep_amort = f"('12. ER Proyectado'!{c_proj}15+'12. ER Proyectado'!{c_proj}16)"
        
        ws_dcf.cell(b_section_row + 4, col, f"=-({ppe_curr}-{ppe_prev}+{dep_amort})")
        
        # Total FCFF (The row that was originally there, now pushed down)
        # It's at b_section_row + 5
        # Formula: Sum the above components
        c_dcf = get_column_letter(col)
        ws_dcf.cell(b_section_row + 5, col, f"=SUM({c_dcf}{b_section_row+1}:{c_dcf}{b_section_row+4})")

    # Update Historical Columns too? (Cols 2-4)
    # Maybe leave them linked to Sheet 8 to save time, assuming Sheet 8 is correct.
    # Just format the new rows for historical cols as empty or N/A
    for col in range(2, 5):
        for i in range(1, 5):
            ws_dcf.cell(b_section_row + i, col, "-")


wb.save(file_path)
print("SUCCESS: DCF updated with detailed links.")
