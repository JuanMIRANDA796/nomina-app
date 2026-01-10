import openpyxl

wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = openpyxl.load_workbook(wb_file) # Open with formulas

print("--- SHEET: 13. BSC Proyectado ---")
ws = wb["13. BSC Proyectado"]

# Check Control Row (Row 1)
print("Control Row (2026-2030):")
for col in range(3, 8):
    cell = ws.cell(row=1, column=col)
    print(f"Col {col} Formula: {cell.value}")

# Check PLUG Row (Row 23 = 18 + 5 offset)
# Map: 'Efectivo y equivalentes de efectivo' -> Row 23
print("\nPlug Formula (Row 23) - 2026:")
print(ws.cell(row=23, column=3).value)

print("\n--- SHEET: 14. WACC Proyectado ---")
ws_wacc = wb["14. WACC Proyectado"]
print("WACC Formula (Row 10):")
print(ws_wacc.cell(row=10, column=2).value)
