import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("="*80)
print("AÑADIENDO ANÁLISIS VERTICAL, HORIZONTAL E INDICADORES")
print("="*80)

# Load existing workbook
wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = load_workbook(wb_file)

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
subheader_fill = PatternFill(start_color="7BA0C9", end_color="7BA0C9", fill_type="solid")
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
formula_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
percent_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
variation_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
bold_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def format_subheader(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = subheader_fill
        cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        cell.alignment = center_align
        cell.border = thin_border

# ==================== SHEET 3: ANÁLISIS VERTICAL - BALANCE ====================
print("\n📊 Creando hoja 3: Análisis Vertical - Balance...")
ws_av_bal = wb.create_sheet("3. Análisis Vertical BSC")
ws_av_bal.sheet_view.showGridLines = False

# Title
ws_av_bal['A1'] = 'ISAGEN S.A. E.S.P.'
ws_av_bal['A1'].font = Font(name='Arial', size=14, bold=True)
ws_av_bal['A2'] = 'ANÁLISIS VERTICAL - BALANCE GENERAL'
ws_av_bal['A2'].font = Font(name='Arial', size=12, bold=True)
ws_av_bal['A3'] = 'Composición porcentual respecto al Total de Activos (%)'
ws_av_bal['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_av_bal.cell(row, 1, 'CONCEPTO')
ws_av_bal.cell(row, 2, '2020')
ws_av_bal.cell(row, 3, '2021')
ws_av_bal.cell(row, 4, '2022')
ws_av_bal.cell(row, 5, '2023')
ws_av_bal.cell(row, 6, '2024')
ws_av_bal.cell(row, 7, '2025 (Sep)')
format_header(ws_av_bal, row, 7)

# Copy concepts from Balance sheet and create formulas
ws_balance = wb["1. Balance General"]
total_activos_row = None

# First pass: find TOTAL ACTIVOS row
for r in range(6, 70):
    concepto = ws_balance.cell(r, 1).value
    if concepto and 'TOTAL ACTIVOS' in str(concepto).upper() and 'CORRIENTES' not in str(concepto).upper():
        total_activos_row = r
        break

print(f"   Total Activos en fila: {total_activos_row}")

# Second pass: create vertical analysis with formulas
current_row = 6
for r in range(6, total_activos_row + 30):  # Include all balance sheet items
    concepto = ws_balance.cell(r, 1).value
    if concepto is None or str(concepto).strip() == '':
        continue
    
    # Copy concepto
    ws_av_bal.cell(current_row, 1, concepto)
    ws_av_bal.cell(current_row, 1).font = normal_font
    ws_av_bal.cell(current_row, 1).border = thin_border
    
    # Create formulas for vertical analysis (each item / Total Activos * 100)
    for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
        formula = f"=('1. Balance General'!{col_letter}{r}/'1. Balance General'!{col_letter}{total_activos_row})*100"
        ws_av_bal.cell(current_row, col_idx, formula)
        ws_av_bal.cell(current_row, col_idx).number_format = '0.00"%"'
        ws_av_bal.cell(current_row, col_idx).alignment = right_align
        ws_av_bal.cell(current_row, col_idx).border = thin_border
        ws_av_bal.cell(current_row, col_idx).fill = percent_fill
    
    # Bold for sections and totals
    if any(x in str(concepto).upper() for x in ['TOTAL', 'ACTIVOS NO CORRIENTES', 'ACTIVOS CORRIENTES', 'PASIVOS', 'PATRIMONIO']):
        for col in range(1, 8):
            ws_av_bal.cell(current_row, col).font = bold_font
            ws_av_bal.cell(current_row, col).fill = total_fill
    
    current_row += 1

# Adjust column widths
ws_av_bal.column_dimensions['A'].width = 50
for col in range(2, 8):
    ws_av_bal.column_dimensions[get_column_letter(col)].width = 14

print("   ✅ Análisis Vertical Balance creado con fórmulas")

# ==================== SHEET 4: ANÁLISIS HORIZONTAL - BALANCE ====================
print("📊 Creando hoja 4: Análisis Horizontal - Balance...")
ws_ah_bal = wb.create_sheet("4. Análisis Horizontal BSC")
ws_ah_bal.sheet_view.showGridLines = False

# Title
ws_ah_bal['A1'] = 'ISAGEN S.A. E.S.P.'
ws_ah_bal['A1'].font = Font(name='Arial', size=14, bold=True)
ws_ah_bal['A2'] = 'ANÁLISIS HORIZONTAL - BALANCE GENERAL'
ws_ah_bal['A2'].font = Font(name='Arial', size=12, bold=True)
ws_ah_bal['A3'] = 'Variación porcentual interanual (%)'
ws_ah_bal['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_ah_bal.cell(row, 1, 'CONCEPTO')
ws_ah_bal.cell(row, 2, '2020-2021')
ws_ah_bal.cell(row, 3, '2021-2022')
ws_ah_bal.cell(row, 4, '2022-2023')
ws_ah_bal.cell(row, 5, '2023-2024')
ws_ah_bal.cell(row, 6, '2024-2025')
format_header(ws_ah_bal, row, 6)

# Create horizontal analysis with formulas
current_row = 6
for r in range(6, total_activos_row + 30):
    concepto = ws_balance.cell(r, 1).value
    if concepto is None or str(concepto).strip() == '':
        continue
    
    # Copy concepto
    ws_ah_bal.cell(current_row, 1, concepto)
    ws_ah_bal.cell(current_row, 1).font = normal_font
    ws_ah_bal.cell(current_row, 1).border = thin_border
    
    # Create formulas for horizontal analysis ((Year2 - Year1) / Year1 * 100)
    col_pairs = [('C', 'B'), ('D', 'C'), ('E', 'D'), ('F', 'E'), ('G', 'F')]
    for col_idx, (year2, year1) in enumerate(col_pairs, start=2):
        formula = f"=(('1. Balance General'!{year2}{r}-'1. Balance General'!{year1}{r})/'1. Balance General'!{year1}{r})*100"
        ws_ah_bal.cell(current_row, col_idx, formula)
        ws_ah_bal.cell(current_row, col_idx).number_format = '0.00"%"'
        ws_ah_bal.cell(current_row, col_idx).alignment = right_align
        ws_ah_bal.cell(current_row, col_idx).border = thin_border
        ws_ah_bal.cell(current_row, col_idx).fill = variation_fill
    
    # Bold for sections and totals
    if any(x in str(concepto).upper() for x in ['TOTAL', 'ACTIVOS NO CORRIENTES', 'ACTIVOS CORRIENTES', 'PASIVOS', 'PATRIMONIO']):
        for col in range(1, 7):
            ws_ah_bal.cell(current_row, col).font = bold_font
    
    current_row += 1

# Adjust column widths
ws_ah_bal.column_dimensions['A'].width = 50
for col in range(2, 7):
    ws_ah_bal.column_dimensions[get_column_letter(col)].width = 14

print("   ✅ Análisis Horizontal Balance creado con fórmulas")

# ==================== SHEET 5: ANÁLISIS VERTICAL - ESTADO RESULTADOS ====================
print("📊 Creando hoja 5: Análisis Vertical - Estado de Resultados...")
ws_av_er = wb.create_sheet("5. Análisis Vertical ER")
ws_av_er.sheet_view.showGridLines = False

# Title
ws_av_er['A1'] = 'ISAGEN S.A. E.S.P.'
ws_av_er['A1'].font = Font(name='Arial', size=14, bold=True)
ws_av_er['A2'] = 'ANÁLISIS VERTICAL - ESTADO DE RESULTADOS'
ws_av_er['A2'].font = Font(name='Arial', size=12, bold=True)
ws_av_er['A3'] = 'Composición porcentual respecto a Ingresos (%)'
ws_av_er['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_av_er.cell(row, 1, 'CONCEPTO')
ws_av_er.cell(row, 2, '2020')
ws_av_er.cell(row, 3, '2021')
ws_av_er.cell(row, 4, '2022')
ws_av_er.cell(row, 5, '2023')
ws_av_er.cell(row, 6, '2024')
ws_av_er.cell(row, 7, '2025 (9m)')
format_header(ws_av_er, row, 7)

# Find Ingresos row
ws_er = wb["2. Estado de Resultados"]
ingresos_row = 6  # First row is Ingresos de contratos con clientes

# Create vertical analysis with formulas
current_row = 6
for r in range(6, 35):  # All Estado de Resultados rows
    concepto = ws_er.cell(r, 1).value
    if concepto is None or str(concepto).strip() == '':
        continue
    
    # Copy concepto
    ws_av_er.cell(current_row, 1, concepto)
    ws_av_er.cell(current_row, 1).font = normal_font
    ws_av_er.cell(current_row, 1).border = thin_border
    
    # Create formulas for vertical analysis (each item / Ingresos * 100)
    for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G'], start=2):
        formula = f"=('2. Estado de Resultados'!{col_letter}{r}/'2. Estado de Resultados'!{col_letter}{ingresos_row})*100"
        ws_av_er.cell(current_row, col_idx, formula)
        ws_av_er.cell(current_row, col_idx).number_format = '0.00"%"'
        ws_av_er.cell(current_row, col_idx).alignment = right_align
        ws_av_er.cell(current_row, col_idx).border = thin_border
        ws_av_er.cell(current_row, col_idx).fill = percent_fill
    
    # Bold for key lines
    if any(x in str(concepto).upper() for x in ['UTILIDAD', 'EBIT', 'EBITDA', 'RESULTADO', 'INTEGRAL', 'BRUTA']):
        for col in range(1, 8):
            ws_av_er.cell(current_row, col).font = bold_font
    
    current_row += 1

# Adjust column widths
ws_av_er.column_dimensions['A'].width = 50
for col in range(2, 8):
    ws_av_er.column_dimensions[get_column_letter(col)].width = 14

print("   ✅ Análisis Vertical Estado de Resultados creado con fórmulas")

# ==================== SHEET 6: ANÁLISIS HORIZONTAL - ESTADO RESULTADOS ====================
print("📊 Creando hoja 6: Análisis Horizontal - Estado de Resultados...")
ws_ah_er = wb.create_sheet("6. Análisis Horizontal ER")
ws_ah_er.sheet_view.showGridLines = False

# Title
ws_ah_er['A1'] = 'ISAGEN S.A. E.S.P.'
ws_ah_er['A1'].font = Font(name='Arial', size=14, bold=True)
ws_ah_er['A2'] = 'ANÁLISIS HORIZONTAL - ESTADO DE RESULTADOS'
ws_ah_er['A2'].font = Font(name='Arial', size=12, bold=True)
ws_ah_er['A3'] = 'Variación porcentual interanual (%)'
ws_ah_er['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_ah_er.cell(row, 1, 'CONCEPTO')
ws_ah_er.cell(row, 2, '2020-2021')
ws_ah_er.cell(row, 3, '2021-2022')
ws_ah_er.cell(row, 4, '2022-2023')
ws_ah_er.cell(row, 5, '2023-2024')
ws_ah_er.cell(row, 6, '2024-2025')
format_header(ws_ah_er, row, 6)

# Create horizontal analysis with formulas
current_row = 6
for r in range(6, 35):
    concepto = ws_er.cell(r, 1).value
    if concepto is None or str(concepto).strip() == '':
        continue
    
    # Copy concepto
    ws_ah_er.cell(current_row, 1, concepto)
    ws_ah_er.cell(current_row, 1).font = normal_font
    ws_ah_er.cell(current_row, 1).border = thin_border
    
    # Create formulas for horizontal analysis
    col_pairs = [('C', 'B'), ('D', 'C'), ('E', 'D'), ('F', 'E'), ('G', 'F')]
    for col_idx, (year2, year1) in enumerate(col_pairs, start=2):
        formula = f"=(('2. Estado de Resultados'!{year2}{r}-'2. Estado de Resultados'!{year1}{r})/'2. Estado de Resultados'!{year1}{r})*100"
        ws_ah_er.cell(current_row, col_idx, formula)
        ws_ah_er.cell(current_row, col_idx).number_format = '0.00"%"'
        ws_ah_er.cell(current_row, col_idx).alignment = right_align
        ws_ah_er.cell(current_row, col_idx).border = thin_border
        ws_ah_er.cell(current_row, col_idx).fill = variation_fill
    
    # Bold for key lines
    if any(x in str(concepto).upper() for x in ['UTILIDAD', 'EBIT', 'EBITDA', 'RESULTADO', 'INTEGRAL', 'BRUTA']):
        for col in range(1, 7):
            ws_ah_er.cell(current_row, col).font = bold_font
    
    current_row += 1

# Adjust column widths
ws_ah_er.column_dimensions['A'].width = 50
for col in range(2, 7):
    ws_ah_er.column_dimensions[get_column_letter(col)].width = 14

print("   ✅ Análisis Horizontal Estado de Resultados creado con fórmulas")

# Save workbook
wb.save(wb_file)
print(f"\n💾 Archivo actualizado con análisis vertical y horizontal")
print("="*80)
