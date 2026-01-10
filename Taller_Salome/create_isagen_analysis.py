import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load source data
source_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\Consolidado_Estados_Financieros_2020_2025.xlsx'
df_esf = pd.read_excel(source_file, sheet_name='ESF')
df_er = pd.read_excel(source_file, sheet_name='ER')

# Create new workbook
wb = Workbook()
wb.remove(wb.active)  # Remove default sheet

# Define styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
subheader_fill = PatternFill(start_color="7BA0C9", end_color="7BA0C9", fill_type="solid")
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
bold_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
center_align = Alignment(horizontal='center', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def format_header(ws, row, cols):
    """Apply header formatting"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

def format_subheader(ws, row, cols):
    """Apply subheader formatting"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = subheader_fill
        cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        cell.alignment = center_align
        cell.border = thin_border

def format_total_row(ws, row, cols):
    """Apply total row formatting"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = total_fill
        cell.font = bold_font
        cell.border = thin_border

# ==================== SHEET 1: BALANCE GENERAL (ESF) ====================
ws_esf = wb.create_sheet("1. Balance General")
ws_esf.sheet_view.showGridLines = False

# Title
ws_esf['A1'] = 'ISAGEN S.A. E.S.P.'
ws_esf['A1'].font = Font(name='Arial', size=14, bold=True)
ws_esf['A2'] = 'ESTADO DE SITUACIÓN FINANCIERA'
ws_esf['A2'].font = Font(name='Arial', size=12, bold=True)
ws_esf['A3'] = 'Valores en millones de pesos colombianos (COP)'
ws_esf['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_esf.cell(row, 1, 'CONCEPTO')
ws_esf.cell(row, 2, '2020')
ws_esf.cell(row, 3, '2021')
ws_esf.cell(row, 4, '2022')
ws_esf.cell(row, 5, '2023')
ws_esf.cell(row, 6, '2024')
ws_esf.cell(row, 7, '2025 (Sep)')
format_header(ws_esf, row, 7)

# Copy data from source
current_row = 6
for idx, row_data in df_esf.iterrows():
    ws_esf.cell(current_row, 1, str(row_data['Concepto']))
    ws_esf.cell(current_row, 2, row_data['2020'])
    ws_esf.cell(current_row, 3, row_data['2021'])
    ws_esf.cell(current_row, 4, row_data['2022'])
    ws_esf.cell(current_row, 5, row_data['2023'])
    ws_esf.cell(current_row, 6, row_data['2024'])
    ws_esf.cell(current_row, 7, row_data['2025-09-01 00:00:00'])
    
    # Format numbers
    for col in range(2, 8):
        cell = ws_esf.cell(current_row, col)
        cell.number_format = '#,##0'
        cell.alignment = right_align
        cell.border = thin_border
    
    # Format concept column
    ws_esf.cell(current_row, 1).border = thin_border
    ws_esf.cell(current_row, 1).font = normal_font
    
    # Bold for section headers and totals
    concept = str(row_data['Concepto']).upper()
    if 'TOTAL' in concept or 'ACTIVOS' == concept or 'PASIVOS' == concept or 'PATRIMONIO' == concept:
        for col in range(1, 8):
            ws_esf.cell(current_row, col).font = bold_font
            ws_esf.cell(current_row, col).fill = total_fill
    
    current_row += 1

# Adjust column widths
ws_esf.column_dimensions['A'].width = 45
for col in range(2, 8):
    ws_esf.column_dimensions[get_column_letter(col)].width = 15

# ==================== SHEET 2: ESTADO DE RESULTADOS ====================
ws_er = wb.create_sheet("2. Estado de Resultados")
ws_er.sheet_view.showGridLines = False

# Title
ws_er['A1'] = 'ISAGEN S.A. E.S.P.'
ws_er['A1'].font = Font(name='Arial', size=14, bold=True)
ws_er['A2'] = 'ESTADO DE RESULTADOS INTEGRAL'
ws_er['A2'].font = Font(name='Arial', size=12, bold=True)
ws_er['A3'] = 'Valores en millones de pesos colombianos (COP)'
ws_er['A3'].font = Font(name='Arial', size=9, italic=True)

# Headers
row = 5
ws_er.cell(row, 1, 'CONCEPTO')
ws_er.cell(row, 2, '2020')
ws_er.cell(row, 3, '2021')
ws_er.cell(row, 4, '2022')
ws_er.cell(row, 5, '2023')
ws_er.cell(row, 6, '2024')
ws_er.cell(row, 7, '2025 (9m)')
format_header(ws_er, row, 7)

# Copy data from source
current_row = 6
for idx, row_data in df_er.iterrows():
    ws_er.cell(current_row, 1, str(row_data['Estado_Resultados']))
    ws_er.cell(current_row, 2, row_data['2020'])
    ws_er.cell(current_row, 3, row_data['2021'])
    ws_er.cell(current_row, 4, row_data['2022'])
    ws_er.cell(current_row, 5, row_data['2023'])
    ws_er.cell(current_row, 6, row_data['2024'])
    ws_er.cell(current_row, 7, row_data['2025 (9 meses)'])
    
    # Format numbers
    for col in range(2, 8):
        cell = ws_er.cell(current_row, col)
        cell.number_format = '#,##0'
        cell.alignment = right_align
        cell.border = thin_border
    
    # Format concept column
    ws_er.cell(current_row, 1).border = thin_border
    ws_er.cell(current_row, 1).font = normal_font
    
    # Bold for key lines
    concept = str(row_data['Estado_Resultados']).upper()
    if any(x in concept for x in ['UTILIDAD', 'EBIT', 'EBITDA', 'RESULTADO', 'TOTAL']):
        for col in range(1, 8):
            ws_er.cell(current_row, col).font = bold_font
            ws_er.cell(current_row, col).fill = total_fill
    
    current_row += 1

# Adjust column widths
ws_er.column_dimensions['A'].width = 45
for col in range(2, 8):
    ws_er.column_dimensions[get_column_letter(col)].width = 15

print("✅ Hojas 1 y 2 creadas: Balance General y Estado de Resultados")
print("Continuando con análisis vertical y horizontal...")

# Save progress
wb.save(r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx')
print("✅ Archivo guardado")
