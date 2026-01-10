from openpyxl import load_workbook

# Load workbook
wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = load_workbook(wb_file, data_only=True)

# Check the full Flujos de Caja sheet
ws_fc = wb["8. Flujos de Caja"]

print("="*80)
print("TODAS LAS FILAS DEL ARCHIVO (después de fila 25)")
print("="*80)

for row in range(25, ws_fc.max_row + 1):
    row_data = []
    for col in range(1, 8):
        value = ws_fc.cell(row, col).value
        if value not in [None, '']:
            row_data.append(f"Col{col}: {value}")
    
    if row_data:
        print(f"Fila {row}: {' | '.join(row_data)}")

print("\n" + "="*80)
