from openpyxl import load_workbook

# Load workbook
wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = load_workbook(wb_file, data_only=True)

# Check the Flujos de Caja sheet
ws_fc = wb["8. Flujos de Caja"]

print("="*80)
print("DATOS AGREGADOS EN LA HOJA '8. Flujos de Caja'")
print("="*80)

# Show all rows to find where the user added the market data
print("\nBuscando datos de mercado...")
for row in range(1, ws_fc.max_row + 1):
    row_data = []
    for col in range(1, 8):
        value = ws_fc.cell(row, col).value
        if value is not None:
            row_data.append(str(value))
    
    if row_data:
        # Check if this row contains market parameters (Rf, Rm, Beta, etc.)
        row_text = ' '.join(row_data)
        if any(keyword in row_text.upper() for keyword in ['RF', 'RM', 'BETA', 'RIESGO', 'TES', 'COLCAP', 'WACC', 'KD', 'KE']):
            print(f"Fila {row}: {' | '.join(row_data)}")

print("\n" + "="*80)
