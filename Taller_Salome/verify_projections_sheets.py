import openpyxl

wb_file = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = openpyxl.load_workbook(wb_file)

print("Sheet Count:", len(wb.sheetnames))
print("Sheet Names:", wb.sheetnames)

if "13. BSC Proyectado" in wb.sheetnames:
    ws = wb["13. BSC Proyectado"]
    print("\n--- Checking Consistency Formula in 13. BSC Proyectado ---")
    # Row 4 is check row
    for col in range(3, 8): # Check a few columns
        cell = ws.cell(row=4, column=col)
        print(f"Col {col} Formula: {cell.value}")

if "14. WACC Proyectado" in wb.sheetnames:
    ws = wb["14. WACC Proyectado"]
    print("\n--- Checking WACC Logic in 14. WACC Proyectado ---")
    # Check WACC row (last row is roughly 25-30)
    for row in range(20, 35):
        cell = ws.cell(row=row, column=1)
        if cell.value == "WACC":
            print(f"Found WACC at Row {row}")
            cell_formula = ws.cell(row=row, column=3).value
            print(f"WACC Formula 2026: {cell_formula}")
