import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("="*80)
print("GENERANDO PROYECCIONES FINANCIERAS 2026-2035")
print("="*80)

# Load workbook
file_path = r'C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\web\Taller_Salome\TALLER07_ISAGEN_VALORACION.xlsx'
wb = openpyxl.load_workbook(file_path)

# Styles
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
subheader_fill = PatternFill(start_color="7BA0C9", end_color="7BA0C9", fill_type="solid")
assumption_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Yellow for inputs
calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Green for calcs
check_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") # Red for checks
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
bold_font = Font(name='Arial', size=10, bold=True)
normal_font = Font(name='Arial', size=10)
center_align = Alignment(horizontal='center', vertical='center')
right_align = Alignment(horizontal='right', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def format_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

# ---------------------------------------------------------
# 1. ESTADO DE RESULTADOS PROYECTADO
# ---------------------------------------------------------
print("Creating Sheet: 12. ER Proyectado")
if "12. ER Proyectado" in wb.sheetnames:
    del wb["12. ER Proyectado"]
ws_er_proj = wb.create_sheet("12. ER Proyectado")
ws_er_proj.sheet_view.showGridLines = False

# Title
ws_er_proj['A1'] = "ESTADO DE RESULTADOS PROYECTADO (2026-2035)"
ws_er_proj['A1'].font = Font(size=14, bold=True)

# Headers
row = 3
ws_er_proj.cell(row, 1, "CONCEPTO")
# Years
for i, year in enumerate(range(2025, 2036)): # 2025 is base
    ws_er_proj.cell(row, i+2, year)
format_header(ws_er_proj, row, 12)

# Data Rows
current_row = 4

# --- Assumptions Section ---
ws_er_proj.cell(current_row, 1, "SUPUESTOS").font = bold_font
current_row += 1

assumptions = [
    ("Crecimiento Ventas (%)", 0.04),
    ("Costo de Ventas (% Ventas)", 0.52),
    ("Gastos Admin (% Ventas)", 0.03),
    ("Otros Gastos (% Ventas)", 0.01),
    ("Tasa Impuestos (%)", 0.35)
]
assump_rows = {}

for name, val in assumptions:
    ws_er_proj.cell(current_row, 1, name)
    # 2025 Base (Manual or linked) - let's just put assumptions from 2026 onwards
    ws_er_proj.cell(current_row, 2, "BASE 2025")
    for col in range(3, 13): # 2026-2035
        ws_er_proj.cell(current_row, col, val)
        ws_er_proj.cell(current_row, col).number_format = '0.0%'
        ws_er_proj.cell(current_row, col).fill = assumption_fill
    assump_rows[name] = current_row
    current_row += 1

current_row += 1 # Spacer

# --- Calculations Section ---
# Revenues
rev_row = current_row
ws_er_proj.cell(rev_row, 1, "Ingresos Operacionales")
# Link 2025 Base from Historical ER (Cell G6, assumed annualize or taken as is? User wants projection. I will verify G6 is 2025 annual or partial. The previous sheet says '2025 (9m)'. I should annualize it.)
# Formula for 2025 Base: '2. Estado de Resultados'!G6 / 9 * 12
ws_er_proj.cell(rev_row, 2, "='2. Estado de Resultados'!G6 / 9 * 12")
ws_er_proj.cell(rev_row, 2).number_format = '#,##0'

# Project 2026+
for col in range(3, 13):
    prev_col = get_column_letter(col-1)
    g_col = get_column_letter(col)
    g_row = assump_rows["Crecimiento Ventas (%)"]
    ws_er_proj.cell(rev_row, col, f"={prev_col}{rev_row}*(1+{g_col}{g_row})")
    ws_er_proj.cell(rev_row, col).number_format = '#,##0'
    ws_er_proj.cell(rev_row, col).fill = calc_fill
current_row += 1

# COGS
cogs_row = current_row
ws_er_proj.cell(cogs_row, 1, "Costo de Ventas")
# 2025 Base
ws_er_proj.cell(cogs_row, 2, "='2. Estado de Resultados'!G7 / 9 * 12")
ws_er_proj.cell(cogs_row, 2).number_format = '#,##0'
# Project
for col in range(3, 13):
    rev_col = get_column_letter(col)
    pct_col = get_column_letter(col)
    pct_row = assump_rows["Costo de Ventas (% Ventas)"]
    ws_er_proj.cell(cogs_row, col, f"=-{rev_col}{rev_row}*{pct_col}{pct_row}") # Negative
    ws_er_proj.cell(cogs_row, col).number_format = '#,##0'
current_row += 1

# Gross Profit
gp_row = current_row
ws_er_proj.cell(gp_row, 1, "Utilidad Bruta").font = bold_font
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_er_proj.cell(gp_row, col, f"={c_let}{rev_row}+{c_let}{cogs_row}")
    ws_er_proj.cell(gp_row, col).number_format = '#,##0'
    ws_er_proj.cell(gp_row, col).font = bold_font
current_row += 1

# Expenses
admin_row = current_row
ws_er_proj.cell(admin_row, 1, "Gastos de Administración")
ws_er_proj.cell(admin_row, 2, "='2. Estado de Resultados'!G10 / 9 * 12")
ws_er_proj.cell(admin_row, 2).number_format = '#,##0'
for col in range(3, 13):
    rev_col = get_column_letter(col)
    pct_col = get_column_letter(col)
    pct_row = assump_rows["Gastos Admin (% Ventas)"]
    ws_er_proj.cell(admin_row, col, f"=-{rev_col}{rev_row}*{pct_col}{pct_row}")
    ws_er_proj.cell(admin_row, col).number_format = '#,##0'
current_row += 1

# EBITDA (Approximation: GP - Admin, ignoring D&A for a moment, D&A is usually inside Cost or Expenses.
# In the historical, EBITDA is given. Let's calculate EBIT then add back D&A? Or calc EBITDA then subtract D&A?
# Let's subtract Expenses to get EBITDA proxy or Operating Income pre-D&A?
# Usually Admin Expenses include Depreciation. Let's project EBITDA Margin?
# User wants full ER. Let's start with EBITDA line or EBIT line.
# Historical: Gross Profit -> Expenses -> EBIT.
# Let's follow that structure.
other_exp_row = current_row
ws_er_proj.cell(other_exp_row, 1, "Otros Gastos")
ws_er_proj.cell(other_exp_row, 2, "='2. Estado de Resultados'!G11 / 9 * 12") # Checking row index...
ws_er_proj.cell(other_exp_row, 2).number_format = '#,##0'
for col in range(3, 13):
    rev_col = get_column_letter(col)
    pct_col = get_column_letter(col)
    pct_row = assump_rows["Otros Gastos (% Ventas)"]
    ws_er_proj.cell(other_exp_row, col, f"=-{rev_col}{rev_row}*{pct_col}{pct_row}")
    ws_er_proj.cell(other_exp_row, col).number_format = '#,##0'
current_row += 1

# EBIT
ebit_row = current_row
ws_er_proj.cell(ebit_row, 1, "Utilidad Operativa (EBIT)").font = bold_font
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_er_proj.cell(ebit_row, col, f"={c_let}{gp_row}+{c_let}{admin_row}+{c_let}{other_exp_row}")
    ws_er_proj.cell(ebit_row, col).number_format = '#,##0'
    ws_er_proj.cell(ebit_row, col).font = bold_font
current_row += 1

# Interest Expense (Will link to Debt in BSC)
# We need to forward reference the Debt row in BSC Proyectado.
# Let's assume Interest Rate (Kd) from WACC sheet.
int_row = current_row
ws_er_proj.cell(int_row, 1, "Gastos Financieros")
# Formula: Previous Year Debt * Kd
# We will fill this formula AFTER creating BSC sheet to know the cell reference.
# Placeholder for now.
current_row += 1

# EBT
ebt_row = current_row
ws_er_proj.cell(ebt_row, 1, "Utilidad Antes de Impuestos").font = bold_font
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_er_proj.cell(ebt_row, col, f"={c_let}{ebit_row}+{c_let}{int_row}")
    ws_er_proj.cell(ebt_row, col).number_format = '#,##0'
    ws_er_proj.cell(ebt_row, col).font = bold_font
current_row += 1

# Taxes
tax_row = current_row
ws_er_proj.cell(tax_row, 1, "Impuesto de Renta")
for col in range(2, 13):
    ebt_col = get_column_letter(col)
    rate_col = get_column_letter(col)
    rate_row = assump_rows["Tasa Impuestos (%)"]
    ws_er_proj.cell(tax_row, col, f"=IF({ebt_col}{ebt_row}>0, -{ebt_col}{ebt_row}*{rate_col}{rate_row}, 0)")
    ws_er_proj.cell(tax_row, col).number_format = '#,##0'
current_row += 1

# Net Income
ni_row = current_row
ws_er_proj.cell(ni_row, 1, "Utilidad Neta").font = bold_font
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_er_proj.cell(ni_row, col, f"={c_let}{ebt_row}+{c_let}{tax_row}")
    ws_er_proj.cell(ni_row, col).number_format = '#,##0'
    ws_er_proj.cell(ni_row, col).font = bold_font
    ws_er_proj.cell(ni_row, col).fill = calc_fill


# ---------------------------------------------------------
# 2. BALANCE GENERAL PROYECTADO
# ---------------------------------------------------------
print("Creating Sheet: 13. BSC Proyectado")
if "13. BSC Proyectado" in wb.sheetnames:
    del wb["13. BSC Proyectado"]
ws_bsc_proj = wb.create_sheet("13. BSC Proyectado")
ws_bsc_proj.sheet_view.showGridLines = False

# Title
ws_bsc_proj['A1'] = "BALANCE GENERAL PROYECTADO (2026-2035)"
ws_bsc_proj['A1'].font = Font(size=14, bold=True)

# Headers
row = 3
ws_bsc_proj.cell(row, 1, "CONCEPTO")
for i, year in enumerate(range(2025, 2036)):
    ws_bsc_proj.cell(row, i+2, year)
format_header(ws_bsc_proj, row, 12)

current_row = 4

# --- Consistency Check ---
check_row = current_row
ws_bsc_proj.cell(check_row, 1, "CONTROL (Activo - Pasivo - Patrimonio)").font = bold_font
ws_bsc_proj.cell(check_row, 1).fill = check_fill
# Formula will be added at end of calculations
current_row += 2

# --- Assumptions ---
ws_bsc_proj.cell(current_row, 1, "SUPUESTOS").font = bold_font
current_row += 1
bsc_assumptions = [
    ("Días Cartera (CxC)", 60),
    ("Días Inventario", 40),
    ("Días Proveedores", 45),
    ("CAPEX (% Ventas)", 0.05),
    ("Depreciación (% PPE previo)", 0.05)
]
bsc_assump_rows = {}
for name, val in bsc_assumptions:
    ws_bsc_proj.cell(current_row, 1, name)
    ws_bsc_proj.cell(current_row, 2, "BASE 2025")
    for col in range(3, 13):
        ws_bsc_proj.cell(current_row, col, val)
        ws_bsc_proj.cell(current_row, col).fill = assumption_fill
    bsc_assump_rows[name] = current_row
    current_row += 1
current_row += 1

# --- ACTIVO ---
ws_bsc_proj.cell(current_row, 1, "ACTIVOS").font = bold_font
current_row += 1

# Cash (PLUG)
cash_row = current_row
ws_bsc_proj.cell(cash_row, 1, "Efectivo y Equivalentes (PLUG)").font = Font(bold=True, italic=True)
ws_bsc_proj.cell(cash_row, 2, "='1. Balance General'!G24") # Base 2025
ws_bsc_proj.cell(cash_row, 2).number_format = '#,##0'
# Formula for plug depends on Total Liab + Eq - Other Assets. Will fill later.
current_row += 1

# CxC
ar_row = current_row
ws_bsc_proj.cell(ar_row, 1, "Cuentas por Cobrar")
ws_bsc_proj.cell(ar_row, 2, "='1. Balance General'!G18")
ws_bsc_proj.cell(ar_row, 2).number_format = '#,##0'
for col in range(3, 13):
    days_row = bsc_assump_rows["Días Cartera (CxC)"]
    c_let = get_column_letter(col)
    # ER Sales for same year * Days / 360
    # ER Sales is row 'rev_row' in '12. ER Proyectado'
    formula = f"='12. ER Proyectado'!{c_let}{rev_row} * {c_let}{days_row} / 360"
    ws_bsc_proj.cell(ar_row, col, formula)
    ws_bsc_proj.cell(ar_row, col).number_format = '#,##0'
current_row += 1

# Inventory
inv_row = current_row
ws_bsc_proj.cell(inv_row, 1, "Inventarios")
ws_bsc_proj.cell(inv_row, 2, "='1. Balance General'!G16")
ws_bsc_proj.cell(inv_row, 2).number_format = '#,##0'
for col in range(3, 13):
    days_row = bsc_assump_rows["Días Inventario"]
    c_let = get_column_letter(col)
    # ER COGS (positive) * Days / 360. COGS in ER is negative.
    formula = f"=-'12. ER Proyectado'!{c_let}{cogs_row} * {c_let}{days_row} / 360"
    ws_bsc_proj.cell(inv_row, col, formula)
    ws_bsc_proj.cell(inv_row, col).number_format = '#,##0'
current_row += 1

# Other Current Assets (Keep Constant)
other_ca_row = current_row
ws_bsc_proj.cell(other_ca_row, 1, "Otros Activos Corrientes (Constante)")
ws_bsc_proj.cell(other_ca_row, 2, "='1. Balance General'!G21 - '1. Balance General'!G16 - '1. Balance General'!G18 - '1. Balance General'!G24") # Approx residual
ws_bsc_proj.cell(other_ca_row, 2).number_format = '#,##0'
for col in range(3, 13):
    prev = get_column_letter(col-1)
    ws_bsc_proj.cell(other_ca_row, col, f"={prev}{other_ca_row}")
    ws_bsc_proj.cell(other_ca_row, col).number_format = '#,##0'
current_row += 1

# Total Current Assets
tca_row = current_row
ws_bsc_proj.cell(tca_row, 1, "Total Activos Corrientes").font = bold_font
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_bsc_proj.cell(tca_row, col, f"=SUM({c_let}{cash_row}:{c_let}{other_ca_row})")
    ws_bsc_proj.cell(tca_row, col).number_format = '#,##0'
    ws_bsc_proj.cell(tca_row, col).font = bold_font
    ws_bsc_proj.cell(tca_row, col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
current_row += 2

# Fixed Assets (PPE)
ppe_row = current_row
ws_bsc_proj.cell(ppe_row, 1, "Propiedad Planta y Equipo (Neto)")
ws_bsc_proj.cell(ppe_row, 2, "='1. Balance General'!G6")
ws_bsc_proj.cell(ppe_row, 2).number_format = '#,##0'
for col in range(3, 13):
    c_let = get_column_letter(col)
    prev = get_column_letter(col-1)
    capex_pct = bsc_assump_rows["CAPEX (% Ventas)"]
    dep_pct = bsc_assump_rows["Depreciación (% PPE previo)"]
    
    # PPE = Prev + Capex - Dep
    # Capex = Sales * %
    # Dep = Prev * %
    capex_form = f"('12. ER Proyectado'!{c_let}{rev_row}*{c_let}{capex_pct})"
    dep_form = f"({prev}{ppe_row}*{c_let}{dep_pct})"
    
    ws_bsc_proj.cell(ppe_row, col, f"={prev}{ppe_row} + {capex_form} - {dep_form}")
    ws_bsc_proj.cell(ppe_row, col).number_format = '#,##0'
current_row += 1

# Other Non-Current (Constant)
other_nca_row = current_row
ws_bsc_proj.cell(other_nca_row, 1, "Otros Activos No Corrientes")
ws_bsc_proj.cell(other_nca_row, 2, "='1. Balance General'!G13 + '1. Balance General'!G7 + '1. Balance General'!G8 + '1. Balance General'!G9") # Aggregated
ws_bsc_proj.cell(other_nca_row, 2).number_format = '#,##0'
for col in range(3, 13):
    prev = get_column_letter(col-1)
    ws_bsc_proj.cell(other_nca_row, col, f"={prev}{other_nca_row}")
    ws_bsc_proj.cell(other_nca_row, col).number_format = '#,##0'
current_row += 1

# Total Assets
ta_row = current_row
ws_bsc_proj.cell(ta_row, 1, "TOTAL ACTIVOS").font = bold_font
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_bsc_proj.cell(ta_row, col, f"={c_let}{tca_row} + {c_let}{ppe_row} + {c_let}{other_nca_row}")
    ws_bsc_proj.cell(ta_row, col).number_format = '#,##0'
    ws_bsc_proj.cell(ta_row, col).font = bold_font
    ws_bsc_proj.cell(ta_row, col).fill = total_fill
current_row += 2

# --- PASIVO Y PATRIMONIO ---
ws_bsc_proj.cell(current_row, 1, "PASIVO").font = bold_font
current_row += 1

# CxP
ap_row = current_row
ws_bsc_proj.cell(ap_row, 1, "Cuentas por Pagar")
ws_bsc_proj.cell(ap_row, 2, "='1. Balance General'!G50")
ws_bsc_proj.cell(ap_row, 2).number_format = '#,##0'
for col in range(3, 13):
    days_row = bsc_assump_rows["Días Proveedores"]
    c_let = get_column_letter(col)
    # ER COGS * Days / 360 (COGS is negative, ensure positive Liab)
    formula = f"=-'12. ER Proyectado'!{c_let}{cogs_row} * {c_let}{days_row} / 360"
    ws_bsc_proj.cell(ap_row, col, formula)
    ws_bsc_proj.cell(ap_row, col).number_format = '#,##0'
current_row += 1

# Financial Debt (Constant Roll-forward)
debt_row = current_row
ws_bsc_proj.cell(debt_row, 1, "Deuda Financiera (Total)")
ws_bsc_proj.cell(debt_row, 2, "='1. Balance General'!G38 + '1. Balance General'!G49")
ws_bsc_proj.cell(debt_row, 2).number_format = '#,##0'
for col in range(3, 13):
    prev = get_column_letter(col-1)
    ws_bsc_proj.cell(debt_row, col, f"={prev}{debt_row}")
    ws_bsc_proj.cell(debt_row, col).number_format = '#,##0'
current_row += 1

# Other Liabilities (Constant)
other_liab_row = current_row
ws_bsc_proj.cell(other_liab_row, 1, "Otros Pasivos")
# Total Liab - Debt - AP (Base 2025)
ws_bsc_proj.cell(other_liab_row, 2, "='1. Balance General'!G55 - '1. Balance General'!G50 - ('1. Balance General'!G38 + '1. Balance General'!G49)")
ws_bsc_proj.cell(other_liab_row, 2).number_format = '#,##0'
for col in range(3, 13):
    prev = get_column_letter(col-1)
    ws_bsc_proj.cell(other_liab_row, col, f"={prev}{other_liab_row}")
    ws_bsc_proj.cell(other_liab_row, col).number_format = '#,##0'
current_row += 1

# Total Liabilities
tl_row = current_row
ws_bsc_proj.cell(tl_row, 1, "Total Pasivos").font = bold_font
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_bsc_proj.cell(tl_row, col, f"=SUM({c_let}{ap_row}:{c_let}{other_liab_row})")
    ws_bsc_proj.cell(tl_row, col).number_format = '#,##0'
    ws_bsc_proj.cell(tl_row, col).font = bold_font
current_row += 2

# --- PATRIMONIO ---
ws_bsc_proj.cell(current_row, 1, "PATRIMONIO").font = bold_font
current_row += 1

equity_row = current_row
ws_bsc_proj.cell(equity_row, 1, "Patrimonio Total")
ws_bsc_proj.cell(equity_row, 2, "='1. Balance General'!G36")
ws_bsc_proj.cell(equity_row, 2).number_format = '#,##0'
for col in range(3, 13):
    prev = get_column_letter(col-1)
    c_let = get_column_letter(col)
    # Prev Equity + Net Income (from ER)
    ni_ref = f"'12. ER Proyectado'!{c_let}{ni_row}"
    ws_bsc_proj.cell(equity_row, col, f"={prev}{equity_row} + {ni_ref}")
    ws_bsc_proj.cell(equity_row, col).number_format = '#,##0'
current_row += 1

# Total Pasivo + Patrimonio
tliab_eq_row = current_row
ws_bsc_proj.cell(tliab_eq_row, 1, "TOTAL PASIVO + PATRIMONIO").font = bold_font
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_bsc_proj.cell(tliab_eq_row, col, f"={c_let}{tl_row} + {c_let}{equity_row}")
    ws_bsc_proj.cell(tliab_eq_row, col).number_format = '#,##0'
    ws_bsc_proj.cell(tliab_eq_row, col).font = bold_font
    ws_bsc_proj.cell(tliab_eq_row, col).fill = total_fill

# --- FIX CASH PLUG ---
# Formula: Cash = (Liab + Equity IF cash was 0) - (Assets IF cash was 0)
# Actually: Cash = (Pasivos_sin_cash* + Equity) - (Activos_sin_cash) -- This logic is flawed.
# Balancing: Total Assets MUST EQUAL Liab + Eq.
# Assets = Cash + NonCashAssets
# Liab+Eq = TotalFinancing
# Cash = TotalFinancing - NonCashAssets
for col in range(2, 13):
    c_let = get_column_letter(col)
    # NonCashAssets = Total Assets - Cash. But Total Assets depends on Cash.
    # So sum components: AR + Inv + OtherCA + PPE + OtherNCA
    # Formula = (Total Pasivos + Patrimonio) - (CxC + Inv + OtherCA + PPE + OtherNCA)
    non_cash_sum = f"({c_let}{ar_row}+{c_let}{inv_row}+{c_let}{other_ca_row}+{c_let}{ppe_row}+{c_let}{other_nca_row})"
    total_fin = f"({c_let}{tl_row}+{c_let}{equity_row})"
    ws_bsc_proj.cell(cash_row, col, f"={total_fin} - {non_cash_sum}")

# --- CHECK ROW LOGIC ---
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_bsc_proj.cell(check_row, col, f"={c_let}{ta_row} - {c_let}{tliab_eq_row}")
    ws_bsc_proj.cell(check_row, col).number_format = '#,##0'
    ws_bsc_proj.cell(check_row, col).fill = check_fill


# --- FIX INTEREST EXPENSE IN ER (Circular or Lagged) ---
# Use Previous Year Debt * Kd
# Kd assumption: 10% (Placeholder) or calculated from WACC
kd = 0.10 # 10% default
for col in range(3, 13):
    prev_col_bsc = get_column_letter(col-1)  # Previous year column in BSC (shifted? No, years match)
    # The columns in BSC align with ER columns (Col C = 2026 for both)
    # Debt is in ws_bsc_proj at 'debt_row'
    # Formula: - '13. BSC Proyectado'!PrevCol{debt_row} * Kd
    ws_er_proj.cell(int_row, col, f"=-'13. BSC Proyectado'!{prev_col_bsc}{debt_row} * {kd}")
    ws_er_proj.cell(int_row, col).number_format = '#,##0'


# ---------------------------------------------------------
# 3. WACC PROYECTADO
# ---------------------------------------------------------
print("Creating Sheet: 14. WACC Proyectado")
if "14. WACC Proyectado" in wb.sheetnames:
    del wb["14. WACC Proyectado"]
ws_wacc_proj = wb.create_sheet("14. WACC Proyectado")
ws_wacc_proj.sheet_view.showGridLines = False

# Title
ws_wacc_proj['A1'] = "WACC PROYECTADO (2026-2035)"
ws_wacc_proj['A1'].font = Font(size=14, bold=True)

# Headers
row = 3
ws_wacc_proj.cell(row, 1, "CONCEPTO")
for i, year in enumerate(range(2025, 2036)):
    ws_wacc_proj.cell(row, i+2, year)
format_header(ws_wacc_proj, row, 12)

current_row = 4

# --- Parameters (Constant) ---
ws_wacc_proj.cell(current_row, 1, "PARÁMETROS CONSTANTES").font = bold_font
current_row += 1

params = [
    ("Rf (Tasa Libre Riesgo)", "='9. Costo de Capital WACC'!G28"), # Link to 2025
    ("Rm (Retorno Mercado)", "='9. Costo de Capital WACC'!G27"),
    ("Beta Desapalancado (Bu)", "='9. Costo de Capital WACC'!G35"),
    ("Riesgo País", "='9. Costo de Capital WACC'!G30"),
    ("Tasa Impuestos", "0.35")
]
param_row_map = {}
for name, ref in params:
    ws_wacc_proj.cell(current_row, 1, name)
    for col in range(2, 13):
        ws_wacc_proj.cell(current_row, col, f"={ref}") # Constant
        ws_wacc_proj.cell(current_row, col).number_format = '0.00%'
    param_row_map[name] = current_row
    current_row += 1
current_row += 1

# --- Capital Structure Annual ---
ws_wacc_proj.cell(current_row, 1, "ESTRUCTURA DE CAPITAL").font = bold_font
current_row += 1

# Debt
d_row_w = current_row
ws_wacc_proj.cell(d_row_w, 1, "Deuda (D)")
# Link to BSC
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_wacc_proj.cell(d_row_w, col, f"='13. BSC Proyectado'!{c_let}{debt_row}")
    ws_wacc_proj.cell(d_row_w, col).number_format = '#,##0'
current_row += 1

# Equity
e_row_w = current_row
ws_wacc_proj.cell(e_row_w, 1, "Patrimonio (E)")
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_wacc_proj.cell(e_row_w, col, f"='13. BSC Proyectado'!{c_let}{equity_row}")
    ws_wacc_proj.cell(e_row_w, col).number_format = '#,##0'
current_row += 1

# Total
tot_row_w = current_row
ws_wacc_proj.cell(tot_row_w, 1, "Total (D+E)")
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_wacc_proj.cell(tot_row_w, col, f"={c_let}{d_row_w}+{c_let}{e_row_w}")
    ws_wacc_proj.cell(tot_row_w, col).number_format = '#,##0'
current_row += 1

# Weight D
wd_row = current_row
ws_wacc_proj.cell(wd_row, 1, "Peso Deuda (Wd)")
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_wacc_proj.cell(wd_row, col, f"={c_let}{d_row_w}/{c_let}{tot_row_w}")
    ws_wacc_proj.cell(wd_row, col).number_format = '0.0%'
current_row += 1

# Weight E
we_row = current_row
ws_wacc_proj.cell(we_row, 1, "Peso Patrimonio (We)")
for col in range(2, 13):
    c_let = get_column_letter(col)
    ws_wacc_proj.cell(we_row, col, f"={c_let}{e_row_w}/{c_let}{tot_row_w}")
    ws_wacc_proj.cell(we_row, col).number_format = '0.0%'
current_row += 2

# --- Costs ---
ws_wacc_proj.cell(current_row, 1, "COSTOS DE CAPITAL").font = bold_font
current_row += 1

# Beta Levered
bl_row_w = current_row
ws_wacc_proj.cell(bl_row_w, 1, "Beta Apalancado (BL)")
for col in range(2, 13):
    c_let = get_column_letter(col)
    bu = f"{c_let}{param_row_map['Beta Desapalancado (Bu)']}"
    tax = f"{c_let}{param_row_map['Tasa Impuestos']}"
    de = f"({c_let}{wd_row}/{c_let}{we_row})"
    ws_wacc_proj.cell(bl_row_w, col, f"={bu}*(1+(1-{tax})*{de})")
    ws_wacc_proj.cell(bl_row_w, col).number_format = '0.00'
current_row += 1

# Ke
ke_row_w = current_row
ws_wacc_proj.cell(ke_row_w, 1, "Costo Patrimonio (Ke)")
for col in range(2, 13):
    c_let = get_column_letter(col)
    rf = f"{c_let}{param_row_map['Rf (Tasa Libre Riesgo)']}"
    rm = f"{c_let}{param_row_map['Rm (Retorno Mercado)']}"
    rp = f"{c_let}{param_row_map['Riesgo País']}"
    beta = f"{c_let}{bl_row_w}"
    ws_wacc_proj.cell(ke_row_w, col, f"={rf} + {beta}*({rm}-{rf}) + {rp}")
    ws_wacc_proj.cell(ke_row_w, col).number_format = '0.00%'
    ws_wacc_proj.cell(ke_row_w, col).fill = calc_fill
current_row += 1

# Kd (Assumed constant or linked)
ws_wacc_proj.cell(current_row, 1, "Costo Deuda (Kd) - Histórico")
kd_ref = "='9. Costo de Capital WACC'!G31" # Kd before tax 2025
for col in range(2, 13):
    ws_wacc_proj.cell(current_row, col, f"={kd_ref}")
    ws_wacc_proj.cell(current_row, col).number_format = '0.00%'
kd_row_w = current_row
current_row += 1

# Kd Net
kdnet_row_w = current_row
ws_wacc_proj.cell(kdnet_row_w, 1, "Costo Deuda Neto (Kd * (1-T))")
for col in range(2, 13):
    c_let = get_column_letter(col)
    tax = f"{c_let}{param_row_map['Tasa Impuestos']}"
    ws_wacc_proj.cell(kdnet_row_w, col, f"={c_let}{kd_row_w}*(1-{tax})")
    ws_wacc_proj.cell(kdnet_row_w, col).number_format = '0.00%'
    ws_wacc_proj.cell(kdnet_row_w, col).fill = calc_fill
current_row += 1

# WACC
wacc_row_w = current_row
ws_wacc_proj.cell(wacc_row_w, 1, "WACC").font = Font(size=12, bold=True)
for col in range(2, 13):
    c_let = get_column_letter(col)
    ke = f"{c_let}{ke_row_w}"
    we = f"{c_let}{we_row}"
    kdnet = f"{c_let}{kdnet_row_w}"
    wd = f"{c_let}{wd_row}"
    ws_wacc_proj.cell(wacc_row_w, col, f"={ke}*{we} + {kdnet}*{wd}")
    ws_wacc_proj.cell(wacc_row_w, col).number_format = '0.00%'
    ws_wacc_proj.cell(wacc_row_w, col).fill = calc_fill
    ws_wacc_proj.cell(wacc_row_w, col).font = bold_font


# Adjust Widths
for ws in [ws_er_proj, ws_bsc_proj, ws_wacc_proj]:
    ws.column_dimensions['A'].width = 35
    for i in range(2, 14):
        ws.column_dimensions[get_column_letter(i)].width = 15

# Save
wb.save(file_path)
print("ARCHIVO GUARDADO EXITOSAMENTE con 3 nuevas hojas.")
