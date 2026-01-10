import pandas as pd
import openpyxl

file_path = r"C:\Users\Juan\.gemini\antigravity\scratch\nomina_colombia\legacy_excel\REGISTRO HORAS  (1).xlsm"

try:
    # Load workbook to read formulas
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    print("--- SHEETS ---")
    print(wb.sheetnames)
    
    if "HOJA_BASE" in wb.sheetnames:
        ws = wb["HOJA_BASE"]
        print("\n--- HOJA_BASE (BE2:BH20) ---")
        # Print a bit of context around the target area too, maybe headers in row 1
        
        # Headers for BE1:BH1
        headers = []
        for col in range(57, 61): # BE is 57
            cell = ws.cell(row=1, column=col)
            headers.append(f"{cell.coordinate}: {cell.value}")
        print("Headers:", headers)

        for row in range(2, 21):
            row_data = []
            for col in range(57, 61): # BE(57) to BH(60) -> actually index is 1-based. 
                                     # BE is 57 (AZ=52, BA=53, BB=54, BC=55, BD=56, BE=57)
                cell = ws.cell(row=row, column=col)
                val = cell.value
                row_data.append(f"{cell.coordinate}={val}")
            print(f"Row {row}: {', '.join(row_data)}")
            
        print("\n--- HOJA_BASE (Absence Logic Check) ---")
        # Check first few columns to see structure
        for row in range(1, 5):
            r = [f"{c.coordinate}={c.value}" for c in ws[row][:20]] # First 20 cols
            print(r)
            
    if "FESTIVOS" in wb.sheetnames:
        print("\n--- FESTIVOS ---")
        ws_fest = wb["FESTIVOS"]
        for row in ws_fest.iter_rows(min_row=1, max_row=10, values_only=True):
            print(row)

except Exception as e:
    print(f"Error: {e}")
